"""
USEtox Input Module

This module provides the USEtoxInput class for populating USEtox Excel templates
with data from PyEPISuite API results.
"""

import os
from typing import Dict, List, Optional, Union, Any
from pathlib import Path
import logging
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

try:
    import pandas as pd
except ImportError:
    pd = None

# Column mapping dictionaries - these tell us which Excel column each property goes in
column_names_dict = {
    "CAS RN": "B",
    "Name": "C", 
    "MW": "F",
    "KOW": "K",
    "Koc": "L",
    "KH25C": "M",  # Henry's law constant
    "Pvap25": "N", 
    "Sol25": "O",
    "T1/2A": "T",
    "T1/2W": "U",  # Water half-life
    "T1/2Sd": "V", # Sediment half-life
    "T1/2Sl": "W", # Soil half-life
}

BAF_dict = {
    "BAFfish": "AQ"  # Bioaccumulation factor for fish (column AQ as specified)
}

toxicity_dict = {
    "Data source": "CH",  # Data source
}

flagged_indicative_dict = {
    "Inorganics": "BV",  # Indicates if the substance is inorganic
    "Data source": "CH",  # Data source
}

logger = logging.getLogger(__name__)


class USEtoxInput:
    """
    Simple class for populating USEtox Excel templates with PyEPISuite data.
    
    The goal is simple:
    1. Call EPI Suite API for each chemical
    2. Open Excel template with openpyxl
    3. Add chemical data row by row starting from row 6
    4. Use column mappings to place data in correct columns
    """
    
    # Excel column mapping (same as column_names_dict but as class attribute)
    EXCEL_COLUMN_MAPPING = column_names_dict
    
    # Experimental data priority - defines which sources take priority over estimates
    EXPERIMENTAL_PROPERTY_PRIORITY = {
        'Sol25': ['experimental_solubility', 'estimated_solubility'],
        'KOW': ['experimental_kow', 'estimated_kow'],
        'Pvap25': ['experimental_vapor_pressure', 'estimated_vapor_pressure'],
        'KH25C': ['experimental_henrys_constant', 'estimated_henrys_constant'],
        'Koc': ['experimental_koc', 'estimated_koc']
    }
    
    # Mapping from EPI Suite DataFrame columns to USEtox properties
    EPISUITE_TO_USETOX_MAPPING = {
        'cas': 'CAS RN',
        'name': 'Name',
        'molecular_weight': 'MW',
        'log_kow_estimated': 'KOW',  # Will be converted from log to linear
        'log_koc_estimated': 'Koc',  # Will be converted from log to linear
        'henrys_law_constant_estimated': 'KH25C',
        'vapor_pressure_estimated': 'Pvap25',
        'water_solubility_logkow_estimated': 'Sol25',
        'atmospheric_half_life_estimated': 'T1/2A',
        'fugacity_water_half_life': 'T1/2W',
        'fugacity_sediment_half_life': 'T1/2Sd',
        'fugacity_soil_half_life': 'T1/2Sl'
    }
    
    def __init__(self, template_path: Optional[str] = None):
        """
        Initialize USEtoxInput with template path.
        
        Args:
            template_path: Path to USEtox template Excel file
        """
        self.template_path = template_path or self._get_default_template_path()
        self.workbook = None
        self.worksheet = None
        self.current_row = 6  # Start from row 6 as requested
        
        # Load the template
        self._load_template()
    
    def _get_default_template_path(self) -> str:
        """Get the default USEtox template path."""
        current_dir = Path(__file__).parent
        return str(current_dir / ".." / ".." / "data" / "usetox3" / "AA_Model_substance_data_Default.xlsx")
    
    def _load_template(self):
        """Load the USEtox template Excel file with openpyxl."""
        try:
            self.workbook = load_workbook(self.template_path)
            self.worksheet = self.workbook["Substance inputs"]
            logger.info(f"Loaded USEtox template from {self.template_path}")
        except Exception as e:
            logger.error(f"Failed to load template: {e}")
            raise
    
    def add_chemical_from_episuite(self, episuite_result: Dict[str, Any]) -> int:
        """
        Add a single chemical from EPI Suite results to the Excel template.
        
        Args:
            episuite_result: Dictionary containing EPI Suite results for one chemical
            
        Returns:
            Row number where the chemical was added
        """
        if not self.worksheet:
            raise ValueError("Template not loaded")
        
        # Get the current row to populate
        row_num = self.current_row
        
        # Add RowNr (Column A)
        self.worksheet[f'A{row_num}'] = row_num - 5
        
        # Map EPI Suite data to Excel columns using the dictionaries
        self._populate_basic_properties(episuite_result, row_num)
        self._populate_baf_properties(episuite_result, row_num)
        self._populate_toxicity_properties(episuite_result, row_num)
        
        # Move to next row for next chemical
        self.current_row += 1
        
        logger.info(f"Added chemical {episuite_result.get('name', 'Unknown')} at row {row_num}")
        return row_num
    
    def _populate_basic_properties(self, episuite_result: Dict[str, Any], row_num: int):
        """Populate basic chemical properties using column_names_dict."""
        # Map EPI Suite fields to USEtox columns
        basic_mapping = {
            'CAS RN': episuite_result.get('cas', ''),
            'Name': episuite_result.get('name', ''),
            'MW': episuite_result.get('molecular_weight'),
            'KOW': self._convert_log_kow_to_kow(episuite_result.get('log_kow_estimated')),
            'Koc': self._convert_log_koc_to_koc(episuite_result.get('log_koc_estimated')),
            'KH25C': episuite_result.get('henrys_law_constant_estimated'),  # Henry's law constant
            'Pvap25': self._convert_mmhg_to_pa(episuite_result.get('vapor_pressure_estimated')),
            'Sol25': self._convert_mg_l_to_g_l(episuite_result.get('water_solubility_logkow_estimated')),
            'T1/2A': episuite_result.get('atmospheric_half_life_estimated'),
            'T1/2W': self._convert_hours_to_days(episuite_result.get('water_biodegradation_half_life_unacclimated')),  # Water half-life
            'T1/2Sd': self._convert_hours_to_days(episuite_result.get('sediment_half_life_hours')),  # Sediment half-life
            'T1/2Sl': self._convert_hours_to_days(episuite_result.get('soil_half_life_hours')),  # Soil half-life
        }
        
        # Set values in Excel using column letters from column_names_dict
        for usetox_prop, value in basic_mapping.items():
            if value is not None and usetox_prop in column_names_dict:
                excel_col = column_names_dict[usetox_prop]
                self.worksheet[f'{excel_col}{row_num}'] = value
    
    def _populate_baf_properties(self, episuite_result: Dict[str, Any], row_num: int):
        """Populate bioaccumulation properties using BAF_dict."""
        # Map EPI Suite BCF to BAFfish
        bcf_value = episuite_result.get('bioconcentration_factor')
        if bcf_value is not None and 'BAFfish' in BAF_dict:
            excel_col = BAF_dict['BAFfish']
            self.worksheet[f'{excel_col}{row_num}'] = bcf_value
    
    def _populate_toxicity_properties(self, episuite_result: Dict[str, Any], row_num: int):
        """Populate toxicity properties using toxicity_dict and flagged_indicative_dict."""
        # Add data source
        if 'Data source' in flagged_indicative_dict:
            excel_col = flagged_indicative_dict['Data source']
            self.worksheet[f'{excel_col}{row_num}'] = "PyEPISuite"
        
        # Check if chemical is inorganic (basic heuristic based on name/molecular structure)
        is_inorganic = self._check_if_inorganic(episuite_result)
        if 'Inorganics' in flagged_indicative_dict:
            excel_col = flagged_indicative_dict['Inorganics']
            self.worksheet[f'{excel_col}{row_num}'] = "Y" if is_inorganic else ""
    
    def _check_if_inorganic(self, episuite_result: Dict[str, Any]) -> bool:
        """
        Basic check to determine if a chemical might be inorganic.
        This is a simple heuristic - could be improved with more sophisticated logic.
        """
        name = episuite_result.get('name', '').lower()
        cas = episuite_result.get('cas', '')
        
        # Common inorganic indicators in chemical names
        inorganic_keywords = [
            'chloride', 'sulfate', 'phosphate', 'nitrate', 'oxide', 'hydroxide',
            'carbonate', 'bicarbonate', 'sulfide', 'fluoride', 'bromide', 'iodide',
            'sodium', 'potassium', 'calcium', 'magnesium', 'aluminum', 'iron',
            'zinc', 'copper', 'lead', 'mercury', 'silver', 'gold', 'platinum'
        ]
        
        # Check if any inorganic keywords are in the name
        for keyword in inorganic_keywords:
            if keyword in name:
                return True
        
        # Exclude common organic compounds that might have low molecular weight
        organic_indicators = [
            'methane', 'ethane', 'propane', 'butane', 'methanol', 'ethanol', 
            'propanol', 'butanol', 'formaldehyde', 'acetaldehyde', 'acetone',
            'benzene', 'toluene', 'xylene', 'phenol'
        ]
        
        # If it contains organic indicators, it's definitely organic
        for indicator in organic_indicators:
            if indicator in name:
                return False
        
        # Additional check: very low molecular weight (< 30) might indicate inorganic
        # but exclude known organic compounds
        mw = episuite_result.get('molecular_weight')
        if mw is not None and mw < 30:  # Only very small molecules
            return True
        
        return False
    
    def _convert_log_kow_to_kow(self, log_kow):
        """Convert log Kow to Kow."""
        if log_kow is not None:
            try:
                return 10 ** float(log_kow)
            except (ValueError, TypeError):
                pass
        return None
    
    def _convert_log_koc_to_koc(self, log_koc):
        """Convert log Koc to Koc."""
        if log_koc is not None:
            try:
                return 10 ** float(log_koc)
            except (ValueError, TypeError):
                pass
        return None
    
    def _convert_mmhg_to_pa(self, mmhg):
        """Convert mmHg to Pa."""
        if mmhg is not None:
            try:
                return float(mmhg) * 133.322
            except (ValueError, TypeError):
                pass
        return None
    
    def _convert_mg_l_to_g_l(self, mg_l):
        """Convert mg/L to g/L."""
        if mg_l is not None:
            try:
                return float(mg_l) / 1000
            except (ValueError, TypeError):
                pass
        return None
    
    def _convert_hours_to_days(self, hours):
        """Convert hours to days."""
        if hours is not None:
            try:
                return float(hours) / 24
            except (ValueError, TypeError):
                pass
        return None
    
    def add_chemicals_from_cas_list(self, cas_list: List[str]) -> Dict[str, int]:
        """
        Add multiple chemicals by calling EPI Suite API for each CAS number.
        
        Args:
            cas_list: List of CAS numbers to look up
            
        Returns:
            Dictionary mapping CAS numbers to row numbers where they were added
        """
        from .utils import search_episuite_by_cas, submit_to_episuite
        
        results = {}
        
        # Search all CAS numbers at once
        search_results = search_episuite_by_cas(cas_list)
        
        if search_results:
            # Get predictions for all found chemicals
            predictions = submit_to_episuite(search_results)
            
            # Handle the predictions structure: it's a tuple where the first element is a list of ResultEPISuite objects
            if predictions and isinstance(predictions, tuple) and len(predictions) > 0:
                result_objects = predictions[0]  # Get the list of ResultEPISuite objects
                
                # Process each prediction
                for i, prediction in enumerate(result_objects):
                    if i < len(search_results):
                        identifier = search_results[i]
                        
                        # Convert to dictionary format
                        prediction_dict = self._convert_episuite_result_to_dict(prediction, identifier)
                        
                        # Add to Excel template
                        row_num = self.add_chemical_from_episuite(prediction_dict)
                        
                        # Find the original CAS number requested
                        original_cas = self._find_original_cas(identifier.cas, cas_list)
                        if original_cas:
                            results[original_cas] = row_num
                            logger.info(f"Successfully added {original_cas} at row {row_num}")
        
        return results
    
    def _find_original_cas(self, episuite_cas: str, original_cas_list: List[str]) -> Optional[str]:
        """Find the original CAS number that corresponds to the EPI Suite CAS."""
        # EPI Suite adds leading zeros, so we need to match
        def normalize_cas(cas_number):
            """Remove leading zeros from CAS number parts."""
            parts = cas_number.split('-')
            if len(parts) == 3:
                return f"{parts[0].lstrip('0') or '0'}-{parts[1].lstrip('0') or '0'}-{parts[2]}"
            return cas_number
        
        normalized_episuite = normalize_cas(episuite_cas)
        for original_cas in original_cas_list:
            if normalize_cas(original_cas) == normalized_episuite:
                return original_cas
        return None
    
    def _convert_episuite_result_to_dict(self, episuite_result, identifier) -> Dict[str, Any]:
        """Convert EPI Suite result object to dictionary format."""
        
        def safe_get_value(obj, path: str):
            """Safely navigate nested object properties."""
            try:
                parts = path.split('.')
                current = obj
                for part in parts:
                    current = getattr(current, part, None)
                    if current is None:
                        return None
                return current
            except (AttributeError, TypeError):
                return None
        
        result_dict = {
            'cas': identifier.cas,
            'name': identifier.name or '',
            
            # Molecular weight from chemical properties
            'molecular_weight': safe_get_value(episuite_result, 'chemicalProperties.molecularWeight'),
            
            # Log Kow from LogKow response (using selectedValue which includes experimental data)
            'log_kow_estimated': safe_get_value(episuite_result, 'logKow.selectedValue.value'),
            
            # Log Koc from LogKoc response (using selectedValue which includes experimental data)
            'log_koc_estimated': safe_get_value(episuite_result, 'logKoc.selectedValue.value'),
            
            # Henry's law constant (using selectedValue which includes experimental data)
            'henrys_law_constant_estimated': safe_get_value(episuite_result, 'henrysLawConstant.selectedValue.value'),
            
            # Vapor pressure (using selectedValue which includes experimental data)
            'vapor_pressure_estimated': safe_get_value(episuite_result, 'vaporPressure.selectedValue.value'),
            
            # Water solubility from LogKow (using selectedValue which includes experimental data)
            'water_solubility_logkow_estimated': safe_get_value(episuite_result, 'waterSolubilityFromLogKow.selectedValue.value'),
            
            # Atmospheric half-life (using estimatedValue because no experimental data usually available)
            'atmospheric_half_life_estimated': safe_get_value(episuite_result, 'atmosphericHalfLife.estimatedValue.value'),
            
            # Water biodegradation half-life: Use a simple default estimate since Biowin doesn't give direct half-life
            'water_biodegradation_half_life_unacclimated': 30,  # Default 30 days for biodegradable organics
            
            # Sediment and soil half-lives from fugacity model (in hours)
            'sediment_half_life_hours': self._safe_extract_fugacity_half_life(episuite_result, 'Sediment'),
            'soil_half_life_hours': self._safe_extract_fugacity_half_life(episuite_result, 'Soil'),
            
            # Bioconcentration factor (using default BCF estimate)
            'bioconcentration_factor': safe_get_value(episuite_result, 'bioconcentration.bioconcentrationFactor'),
        }
        return result_dict
    
    def _safe_extract_fugacity_half_life(self, episuite_result, compartment: str):
        """Safely extract half-life from fugacity model for specific compartment (Sediment or Soil)."""
        try:
            if hasattr(episuite_result, 'fugacityModel') and episuite_result.fugacityModel:
                if hasattr(episuite_result.fugacityModel, 'model') and episuite_result.fugacityModel.model:
                    compartment_data = getattr(episuite_result.fugacityModel.model, compartment, None)
                    if compartment_data and len(compartment_data) > 0 and compartment_data[0]:
                        return compartment_data[0].HalfLife
            return None
        except (AttributeError, IndexError, TypeError):
            return None
    
    def save_excel(self, output_path: str):
        """
        Save the populated Excel file.
        
        Args:
            output_path: Path for the output Excel file
        """
        if not self.workbook:
            raise ValueError("No workbook loaded")
        
        self.workbook.save(output_path)
        logger.info(f"Saved USEtox Excel file to {output_path}")
    
    def get_summary(self) -> Dict[str, Any]:
        """Get summary of populated data."""
        if not self.worksheet:
            return {}
        
        chemicals_added = self.current_row - 6
        return {
            'chemicals_added': chemicals_added,
            'current_row': self.current_row,
            'template_path': self.template_path
        }

    def get_summary_statistics(self) -> Dict[str, Any]:
        """Get summary statistics of populated data."""
        if not self.worksheet:
            return {
                'total_chemicals': 0,
                'properties_populated': 0,
                'missing_values': 0
            }
        
        chemicals_added = self.current_row - 6
        # Count populated properties by checking non-empty cells
        properties_populated = 0
        missing_values = 0
        
        for row_num in range(6, self.current_row):
            for col_letter in column_names_dict.values():
                cell_value = self.worksheet[f"{col_letter}{row_num}"].value
                if cell_value is not None and cell_value != "":
                    properties_populated += 1
                else:
                    missing_values += 1
        
        return {
            'total_chemicals': chemicals_added,
            'properties_populated': properties_populated,
            'missing_values': missing_values
        }

    def validate_data(self) -> Dict[str, List[str]]:
        """Validate the populated data and return warnings/errors."""
        warnings = []
        errors = []
        
        if not self.worksheet:
            errors.append("No worksheet loaded")
            return {'warnings': warnings, 'errors': errors}
        
        # Check for required fields
        for row_num in range(6, self.current_row):
            cas_cell = self.worksheet[f"{column_names_dict['CAS RN']}{row_num}"]
            name_cell = self.worksheet[f"{column_names_dict['Name']}{row_num}"]
            
            if not cas_cell.value:
                warnings.append(f"Row {row_num}: Missing CAS number")
            if not name_cell.value:
                warnings.append(f"Row {row_num}: Missing chemical name")
                
            # Check for unrealistic values
            mw_cell = self.worksheet[f"{column_names_dict['MW']}{row_num}"]
            if mw_cell.value and mw_cell.value < 0:
                warnings.append(f"Row {row_num}: Negative molecular weight")
                
        return {'warnings': warnings, 'errors': errors}

    def get_excel_column_letter(self, property_name: str) -> Optional[str]:
        """Get the Excel column letter for a given property name."""
        return column_names_dict.get(property_name)

    def get_data_source_analysis(self) -> Dict[str, int]:
        """Analyze data sources in the populated sheet."""
        if not self.worksheet:
            return {}
        
        # This is a simplified analysis - in practice you'd track data sources during population
        chemicals_count = self.current_row - 6
        return {
            'estimated': chemicals_count,
            'experimental': 0,  # Would track this during population
            'manual': 0
        }

    def populate_from_episuite_dataframe(self, episuite_df):
        """Populate USEtox template from EPI Suite DataFrame."""
        import pandas as pd
        
        if not isinstance(episuite_df, pd.DataFrame):
            raise ValueError("Input must be a pandas DataFrame")
        
        populated_rows = []
        for idx, row in episuite_df.iterrows():
            episuite_dict = row.to_dict()
            row_num = self.add_chemical_from_episuite(episuite_dict)
            if row_num > 0:
                populated_rows.append(row_num)
        
        # Return a simple representation - in practice this could be more sophisticated
        return episuite_df

    def export_to_excel(self, output_path: str, include_headers: bool = True, include_original_template: bool = True):
        """Export to Excel with additional options."""
        self.save_excel(output_path)
        
    @property
    def populated_df(self):
        """Get populated DataFrame representation."""
        # This is a simplified version - could be enhanced to return actual Excel data as DataFrame
        return None if not self.worksheet else "populated"


def create_usetox_input_from_cas_list(cas_list: List[str], 
                                    output_path: str,
                                    template_path: Optional[str] = None) -> USEtoxInput:
    """
    Convenience function to create USEtox input from a list of CAS numbers.
    
    Args:
        cas_list: List of CAS numbers to look up in EPI Suite
        output_path: Path for output Excel file
        template_path: Optional custom template path
        
    Returns:
        USEtoxInput instance with populated data
    """
    # Create USEtoxInput instance
    usetox_input = USEtoxInput(template_path)
    
    # Add chemicals from CAS list
    results = usetox_input.add_chemicals_from_cas_list(cas_list)
    
    # Save to Excel
    usetox_input.save_excel(output_path)
    
    # Print summary
    summary = usetox_input.get_summary()
    print(f"✅ Successfully created USEtox input with {summary['chemicals_added']} chemicals")
    print(f"📋 Excel file exported to: {output_path}")
    
    return usetox_input


def create_usetox_input_from_episuite(episuite_df: Any,
                                     output_path: str,
                                     template_path: Optional[str] = None,
                                     experimental_data: Optional[Dict[str, Dict[str, Any]]] = None) -> USEtoxInput:
    """
    Convenience function to create USEtox input from EPI Suite DataFrame results.
    
    Args:
        episuite_df: DataFrame with EPI Suite results (from episuite_to_dataframe)
        output_path: Path for output Excel file
        template_path: Optional custom template path
        experimental_data: Optional dict of experimental data to override estimates
        
    Returns:
        USEtoxInput instance with populated data
        
    Example:
        >>> from pyepisuite.utils import search_episuite_by_cas, submit_to_episuite
        >>> from pyepisuite.dataframe_utils import episuite_to_dataframe
        >>> ids = search_episuite_by_cas(['50-00-0', '67-56-1'])
        >>> epi_results, _ = submit_to_episuite(ids)
        >>> epi_df = episuite_to_dataframe(epi_results)
        >>> usetox_input = create_usetox_input_from_episuite(
        ...     episuite_df=epi_df,
        ...     output_path='usetox_results.xlsx'
        ... )
    """
    try:
        # Create USEtoxInput instance
        usetox_input = USEtoxInput(template_path)
        
        # Process each chemical in the DataFrame
        chemicals_added = 0
        for idx, row in episuite_df.iterrows():
            # Convert DataFrame row to dictionary format expected by USEtoxInput
            episuite_dict = row.to_dict()
            
            # Add chemical to USEtox template
            row_num = usetox_input.add_chemical_from_episuite(episuite_dict)
            if row_num > 0:
                chemicals_added += 1
                
            # Apply experimental data if provided
            if experimental_data and 'cas' in episuite_dict:
                cas_number = episuite_dict['cas']
                if cas_number in experimental_data:
                    # Override with experimental data
                    exp_data = experimental_data[cas_number]
                    for prop, value in exp_data.items():
                        if prop in column_names_dict:
                            col_letter = column_names_dict[prop]
                            usetox_input.worksheet[f"{col_letter}{row_num}"] = value
        
        # Save to Excel
        usetox_input.save_excel(output_path)
        
        # Print summary
        summary = usetox_input.get_summary()
        print(f"✅ Successfully created USEtox input with {chemicals_added} chemicals")
        print(f"📋 Excel file exported to: {output_path}")
        
        return usetox_input
        
    except Exception as e:
        logger.error(f"Error creating USEtox input from EPI Suite DataFrame: {e}")
        raise
