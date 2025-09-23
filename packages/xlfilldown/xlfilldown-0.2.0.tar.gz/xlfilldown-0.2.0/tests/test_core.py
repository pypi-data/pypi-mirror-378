import io
from pathlib import Path
import sqlite3
from datetime import datetime

import openpyxl
import pytest

from xlfilldown.core import qident, normalize_headers, canon_list, sha256_hex
from xlfilldown.api import ingest_excel_to_sqlite, ingest_excel_to_excel

from decimal import Decimal
from xlfilldown.core import canon_list, sha256_hex

import subprocess
import sys
import os
import textwrap
import os
import shutil
from pathlib import Path
import openpyxl
import pytest
from xlfilldown.api import ingest_excel_to_excel


def _make_workbook(tmp_path: Path):
    p = tmp_path / "in.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    # Header at row 1
    ws.append(["tier0", "tier1"])
    # Data rows per the example:
    # value1, tiervalue1
    ws.append(["value1", "tiervalue1"])
    # empty,  tiervalue2
    ws.append([None, "tiervalue2"])
    # empty,  empty   -> completely empty row (no fill-down applied)
    ws.append([None, None])
    # empty,  tiervalue3
    ws.append([None, "tiervalue3"])
    # value2, tiervalue4
    ws.append(["value2", "tiervalue4"])
    # empty,  tiervalue5
    ws.append([None, "tiervalue5"])
    wb.save(p)
    return p


def test_qident():
    assert qident('a') == '"a"'
    assert qident('a"b') == '"a""b"'


def test_normalize_headers():
    headers = normalize_headers([None, ' A ', 'nan', 'ok'])
    assert headers == ['', 'A', '', 'ok']


def test_canon_list_and_hash():
    s = canon_list([None, ' A ', 5])
    assert s == '["","A","5"]'
    h = sha256_hex(s)
    assert len(h) == 64


def test_ingest_to_sqlite_with_empty_row_handling(tmp_path):
    infile = _make_workbook(tmp_path)
    db_path = tmp_path / "out.db"

    summary = ingest_excel_to_sqlite(
        file=infile,
        sheet="Sheet1",
        header_row=1,
        pad_cols=["tier0"],
        db=db_path,
        table="t",
        drop_blank_rows=False,
        require_non_null=[],
        row_hash=True,
        excel_row_numbers=True,
        if_exists="fail",
        batch_size=50,
    )

    assert summary["table"] == "t"
    assert summary["row_hash"] is True
    assert summary["excel_row_numbers"] is True

    # Verify rows and the "completely empty" row behavior: row 4 should remain empty
    with sqlite3.connect(str(db_path)) as conn:
        cur = conn.cursor()
        rows = cur.execute('SELECT excel_row, tier0, tier1 FROM t ORDER BY excel_row').fetchall()
        # Expect 6 data rows (we have 6 input data rows after header)
        assert len(rows) == 6
        # Row mapping by excel_row (2..7)
        by_row = {r[0]: (r[1], r[2]) for r in rows}
        assert by_row["2"] == ("value1", "tiervalue1")
        assert by_row["3"] == ("value1", "tiervalue2")  # filled down
        assert by_row["4"] == (None, None)  # completely empty row preserved
        assert by_row["5"] == ("value1", "tiervalue3")  # carry persists past empty row
        assert by_row["6"] == ("value2", "tiervalue4")
        assert by_row["7"] == ("value2", "tiervalue5")  # filled down


def test_ingest_to_excel_with_append_and_headers(tmp_path):
    infile = _make_workbook(tmp_path)
    outfile = tmp_path / "out.xlsx"

    # First write (new workbook, new sheet)
    summary1 = ingest_excel_to_excel(
        file=infile,
        sheet="Sheet1",
        header_row=1,
        pad_cols=["tier0"],
        outfile=outfile,
        outsheet="Processed",
        drop_blank_rows=False,
        require_non_null=[],
        row_hash=False,
        excel_row_numbers=False,
        if_exists="fail",
    )
    assert summary1["sheet"] == "Processed"
    assert summary1["rows_written"] == 6

    # Second write append; must match headers
    summary2 = ingest_excel_to_excel(
        file=infile,
        sheet="Sheet1",
        header_row=1,
        pad_cols=["tier0"],
        outfile=outfile,
        outsheet="Processed",
        drop_blank_rows=False,
        require_non_null=[],
        row_hash=False,
        excel_row_numbers=False,
        if_exists="append",
    )
    assert summary2["rows_written"] == 6

    # Validate workbook content: header + 12 data rows
    wb = openpyxl.load_workbook(outfile)
    ws = wb["Processed"]
    assert ws.max_row == 1 + 12  # header + rows
    # Check a couple of expected values
    assert ws.cell(row=2, column=1).value == "value1"
    assert ws.cell(row=3, column=1).value == "value1"
    # completely empty row at row 4 in the first block -> row 5 in sheet (header row=1)
    assert ws.cell(row=4, column=1).value is None and ws.cell(row=4, column=2).value is None


def test_row_hash_float_decimal_stability():
    """1, 1.0, and Decimal('1.00') should canonicalize the same for hashing."""
    cases = [
        [1, 2.50, 1000000],
        [1.0, 2.5, 1_000_000.0],
        [Decimal("1"), Decimal("2.5000"), Decimal("1000000.000")],
    ]
    texts = [canon_list(c) for c in cases]
    hashes = [sha256_hex(t) for t in texts]
    # All canon_list strings should be identical (or at least the hashes)
    assert len(set(texts)) == 1
    assert len(set(hashes)) == 1


def test_whitespace_only_cells_are_blank_for_padding_and_filters(tmp_path):
    """Whitespace-only cells should behave like blank for padding and require_non_null."""
    p = tmp_path / "in_ws.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["A", "B"])
    ws.append(["x", "   "])   # B is whitespace-only -> blank
    ws.append([None, "y"])    # should fill A down to "x"
    wb.save(p)

    db_path = tmp_path / "o.db"
    summary = ingest_excel_to_sqlite(
        file=p,
        sheet="S",
        header_row=1,
        pad_cols=["A"],
        db=db_path,
        table="t",
        drop_blank_rows=False,
        require_non_null=["A"],  # enforce non-null after padding
        row_hash=False,
        excel_row_numbers=False,
        if_exists="fail",
        batch_size=10,
    )
    assert summary["rows_ingested"] == 2

    with sqlite3.connect(str(db_path)) as conn:
        rows = conn.execute("SELECT A, B FROM t ORDER BY rowid").fetchall()
    # Row 1: B becomes None (treated blank); Row 2: A filled down to 'x'
    assert rows == [("x", None), ("x", "y")]


def test_drop_blank_rows_with_whitespace_only_pad_cols(tmp_path):
    """Rows blank across pad_cols should drop even if cells contain only whitespace."""
    p = tmp_path / "in_drop.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["grp", "val"])
    ws.append([None, "  "])   # blank across pad_cols=['grp'] -> droppable
    ws.append(["", "   "])    # also droppable
    ws.append(["g1", "v1"])   # keep
    wb.save(p)

    db_path = tmp_path / "o2.db"
    summary = ingest_excel_to_sqlite(
        file=p,
        sheet="S",
        header_row=1,
        pad_cols=["grp"],
        db=db_path,
        table="t",
        drop_blank_rows=True,
        require_non_null=[],
        row_hash=False,
        excel_row_numbers=False,
        if_exists="fail",
        batch_size=10,
    )
    assert summary["rows_ingested"] == 1
    with sqlite3.connect(str(db_path)) as conn:
        rows = conn.execute("SELECT grp, val FROM t").fetchall()
    assert rows == [("g1", "v1")]


def test_row_hash_matches_expected_with_float_padding(tmp_path):
    """
    Verify stored row_hash equals sha256 over canon_list of row values,
    and that floats normalize (1.0 -> '1') even after fill-down.
    """
    p = tmp_path / "in_hash.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["n", "txt"])
    ws.append([1.0, "a"])   # numeric as float
    ws.append([None, "b"])  # will fill n down to 1.0
    wb.save(p)

    db_path = tmp_path / "hash.db"
    summary = ingest_excel_to_sqlite(
        file=p,
        sheet="S",
        header_row=1,
        pad_cols=["n"],
        db=db_path,
        table="t",
        drop_blank_rows=False,
        require_non_null=[],
        row_hash=True,
        excel_row_numbers=False,
        if_exists="fail",
        batch_size=10,
    )
    assert summary["rows_ingested"] == 2

    with sqlite3.connect(str(db_path)) as conn:
        rows = conn.execute("SELECT row_hash, n, txt FROM t ORDER BY rowid").fetchall()

    # Build expected hashes using canon_list (which should normalize 1.0 -> '1')
    expected1 = sha256_hex(canon_list([1, "a"]))  # 1.0 canonical as '1'
    expected2 = sha256_hex(canon_list([1, "b"]))  # filled down 1.0 -> '1'

    assert rows[0][0] == expected1
    assert rows[1][0] == expected2
    # And make sure DB kept numeric value as-is (could be 1.0), but hash stability is what matters.
    assert rows[0][2] == "a" and rows[1][2] == "b"


def test_xlsx_headers_include_row_hash_and_excel_row_in_order(tmp_path):
    """When writing to Excel with both flags, header order should be row_hash, excel_row, then data headers."""
    infile = _make_workbook(tmp_path)
    outfile = tmp_path / "out_hash.xlsx"

    summary = ingest_excel_to_excel(
        file=infile,
        sheet="Sheet1",
        header_row=1,
        pad_cols=["tier0"],
        outfile=outfile,
        outsheet="Processed",
        drop_blank_rows=False,
        require_non_null=[],
        row_hash=True,
        excel_row_numbers=True,
        if_exists="replace",
    )
    assert summary["sheet"] == "Processed"

    wb = openpyxl.load_workbook(outfile)
    ws = wb["Processed"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert header == ["row_hash", "excel_row", "tier0", "tier1"]

def _make_tiered_workbook_for_pad_modes(tmp_path: Path):
    """
    Build a sheet that highlights the behavioral difference between hierarchical and independent fill:
      Row 2: apple group header
      Row 3: red/sour/value1 (fills apple block)
      Row 4: potato group header
      Row 5: value2 (Tier 4 present BEFORE new Tier 2/3 appear)
      Row 6: fried/yellow/value3 (now Tier 2/3 show for potato)
    """
    p = tmp_path / "padmodes.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["Tier 1", "Tier 2", "Tier 3", "Tier 4"])  # header row = 1
    ws.append(["apple",  None,   None,   None])          # 2: group header
    ws.append([None,     "red",  "sour", "value1"])      # 3: detail
    ws.append(["potato", None,   None,   None])          # 4: new group header
    ws.append([None,     None,   None,   "value2"])      # 5: Tier4 BEFORE T2/T3 (key row)
    ws.append([None,     "fried","yellow","value3"])     # 6: detail w/ T2/T3 present
    wb.save(p)
    return p


def test_pad_mode_hierarchical_resets_lower_tiers_with_tier4(tmp_path):
    """
    With hierarchical fill (default), when Tier 1 changes to 'potato', lower-tier carries reset.
    The first 'potato' detail row (row 5) should have Tier 2/3 = None, Tier 4 present.
    Group headers are dropped by require_non_null=['Tier 4'].
    """
    p = _make_tiered_workbook_for_pad_modes(tmp_path)
    db = tmp_path / "hier.db"

    summary = ingest_excel_to_sqlite(
        file=p,
        sheet="S",
        header_row=1,
        pad_cols=["Tier 1", "Tier 2", "Tier 3"],
        db=db,
        table="t",
        drop_blank_rows=True,              # drop true spacer rows only
        require_non_null=["Tier 4"],       # drop group headers
        row_hash=False,
        excel_row_numbers=False,
        if_exists="fail",
        batch_size=50,
        pad_hierarchical=True,             # <--- hierarchical
    )
    assert summary["rows_ingested"] == 3  # value1, value2, value3

    with sqlite3.connect(str(db)) as conn:
        rows = conn.execute('SELECT "Tier 1","Tier 2","Tier 3","Tier 4" FROM t ORDER BY rowid').fetchall()

    # Expect:
    # row for value1 (apple/red/sour)
    # row for value2 (potato/None/None)  <-- lower tiers reset
    # row for value3 (potato/fried/yellow)
    assert rows == [
        ("apple",  "red",   "sour",   "value1"),
        ("potato", None,    None,     "value2"),
        ("potato", "fried", "yellow", "value3"),
    ]


def test_pad_mode_independent_carries_lower_tiers_across_groups(tmp_path):
    """
    With independent (legacy) fill, lower tiers carry over when Tier 1 changes.
    The first 'potato' detail row (row 5) inherits Tier 2/3 = red/sour.
    Group headers are dropped by require_non_null=['Tier 4'].
    """
    p = _make_tiered_workbook_for_pad_modes(tmp_path)
    db = tmp_path / "indep.db"

    summary = ingest_excel_to_sqlite(
        file=p,
        sheet="S",
        header_row=1,
        pad_cols=["Tier 1", "Tier 2", "Tier 3"],
        db=db,
        table="t",
        drop_blank_rows=True,
        require_non_null=["Tier 4"],
        row_hash=False,
        excel_row_numbers=False,
        if_exists="fail",
        batch_size=50,
        pad_hierarchical=False,            # <--- independent/legacy
    )
    assert summary["rows_ingested"] == 3  # value1, value2, value3

    with sqlite3.connect(str(db)) as conn:
        rows = conn.execute('SELECT "Tier 1","Tier 2","Tier 3","Tier 4" FROM t ORDER BY rowid').fetchall()

    # Expect independent carry:
    # row for value1 (apple/red/sour)
    # row for value2 (potato/red/sour)  <-- carried across Tier 1 change
    # row for value3 (potato/fried/yellow)
    assert rows == [
        ("apple",  "red",   "sour",   "value1"),
        ("potato", "red",   "sour",   "value2"),
        ("potato", "fried", "yellow", "value3"),
    ]

def test_duplicate_headers_raise():
    """Duplicate header names after normalization should error."""
    import openpyxl
    from xlfilldown.api import ingest_excel_to_sqlite
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["A", "A"])  # duplicate headers
    ws.append(["x", "y"])
    p = Path(__file__).parent / "dup.xlsx"
    wb.save(p)
    try:
        with pytest.raises(ValueError, match="Duplicate header names"):
            ingest_excel_to_sqlite(
                file=p, sheet="S", header_row=1, pad_cols=["A"], db=Path(__file__).parent / "dup.db"
            )
    finally:
        if p.exists():
            p.unlink(missing_ok=True)
        (Path(__file__).parent / "dup.db").unlink(missing_ok=True)


def test_missing_require_non_null_header_errors(tmp_path):
    """If a header named in require_non_null is not present, raise a clear error."""
    p = tmp_path / "in.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["A", "B"])
    ws.append(["x", "y"])
    wb.save(p)

    with pytest.raises(ValueError, match="require_non_null header\\(s\\) not found"):
        ingest_excel_to_sqlite(
            file=p,
            sheet="S",
            header_row=1,
            pad_cols=["A"],
            db=tmp_path / "o.db",
            require_non_null=["C"],  # not present
        )


def test_header_row_not_one(tmp_path):
    """Header row can be offset; excel_row should reflect original row numbers as TEXT."""
    p = tmp_path / "in_hdr3.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append([None, None])         # row 1 spacer
    ws.append([None, None])         # row 2 spacer
    ws.append(["H1", "H2"])         # row 3 header
    ws.append(["v1", None])         # row 4
    ws.append([None, "v2"])         # row 5
    wb.save(p)

    db = tmp_path / "o_hdr3.db"
    summary = ingest_excel_to_sqlite(
        file=p, sheet="S", header_row=3, pad_cols=["H1"], db=db,
        row_hash=False, excel_row_numbers=True
    )
    assert summary["rows_ingested"] == 2
    with sqlite3.connect(str(db)) as conn:
        rows = conn.execute('SELECT excel_row, "H1", "H2" FROM "S" ORDER BY rowid').fetchall()
    # excel_row is TEXT and should be "4", "5"
    assert rows == [("4", "v1", None), ("5", "v1", "v2")]


def test_sqlite_excel_row_is_text_type(tmp_path):
    """The sqlite column type for excel_row should be TEXT, not INTEGER."""
    p = tmp_path / "in_text.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["A"])
    ws.append(["x"])
    wb.save(p)

    db = tmp_path / "t.db"
    ingest_excel_to_sqlite(
        file=p, sheet="S", header_row=1, pad_cols=["A"], db=db,
        excel_row_numbers=True
    )
    with sqlite3.connect(str(db)) as conn:
        info = conn.execute('PRAGMA table_info("S");').fetchall()
    # columns in order: excel_row (TEXT), A (TEXT)
    names_and_types = [(r[1], r[2]) for r in info]
    assert names_and_types[0] == ("excel_row", "TEXT")
    assert names_and_types[1] == ("A", "TEXT")


def test_xlsx_append_header_mismatch_raises(tmp_path):
    """Appending with different optional columns (e.g., adding row_hash) should fail."""
    infile = _make_workbook(tmp_path)
    outfile = tmp_path / "out_append.xlsx"

    # initial write WITHOUT row_hash/excel_row
    ingest_excel_to_excel(
        file=infile,
        sheet="Sheet1",
        header_row=1,
        pad_cols=["tier0"],
        outfile=outfile,
        outsheet="Processed",
        row_hash=False,
        excel_row_numbers=False,
        if_exists="fail",
    )

    # now try to append WITH row_hash -> header mismatch
    with pytest.raises(ValueError, match="does not match expected"):
        ingest_excel_to_excel(
            file=infile,
            sheet="Sheet1",
            header_row=1,
            pad_cols=["tier0"],
            outfile=outfile,
            outsheet="Processed",
            row_hash=True,                  # <-- new
            excel_row_numbers=False,
            if_exists="append",
        )


def test_sqlite_append_schema_mismatch_raises(tmp_path):
    """Appending with a different schema (e.g., adding excel_row) should fail."""
    infile = _make_workbook(tmp_path)
    db = tmp_path / "out.db"

    # first load WITHOUT excel_row
    ingest_excel_to_sqlite(
        file=infile,
        sheet="Sheet1",
        header_row=1,
        pad_cols=["tier0"],
        db=db,
        table="t",
        excel_row_numbers=False,
        if_exists="replace",
    )

    # now try to append WITH excel_row -> expect mismatch
    with pytest.raises(ValueError, match="does not match expected"):
        ingest_excel_to_sqlite(
            file=infile,
            sheet="Sheet1",
            header_row=1,
            pad_cols=["tier0"],
            db=db,
            table="t",
            excel_row_numbers=True,      # <-- new column
            if_exists="append",
        )


def test_drop_blank_rows_true_vs_false(tmp_path):
    """Whitespace-only rows across pad_cols should be dropped when requested."""
    p = tmp_path / "in_wsdrop.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["grp", "val"])
    ws.append(["   ", "   "])   # whitespace-only -> blank across pad_cols
    ws.append([None, None])     # completely empty
    ws.append(["g1", "x"])
    wb.save(p)

    # drop_blank_rows = False -> we keep the spacer row (and the empty) as empty lines
    db1 = tmp_path / "keep.db"
    s1 = ingest_excel_to_sqlite(
        file=p, sheet="S", header_row=1, pad_cols=["grp"], db=db1,
        drop_blank_rows=False
    )
    assert s1["rows_ingested"] == 3  # spacer, empty, g1

    # drop_blank_rows = True -> drop rows blank across pad_cols; completely empty row also dropped
    db2 = tmp_path / "drop.db"
    s2 = ingest_excel_to_sqlite(
        file=p, sheet="S", header_row=1, pad_cols=["grp"], db=db2,
        drop_blank_rows=True
    )
    assert s2["rows_ingested"] == 1  # only the 'g1' row remains


def test_canonicalization_negative_zero_and_large_numbers(tmp_path):
    """-0.0 should canonicalize to '0'; large numbers should avoid sci-notation for hashing and storage."""
    p = tmp_path / "canon.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["n1", "n2"])
    ws.append([-0.0, 1000000.0])     # -0.0, large float
    ws.append([None, 1000000])       # will fill n1 down to -0.0 (becomes '0')
    wb.save(p)

    db = tmp_path / "canon.db"
    ingest_excel_to_sqlite(
        file=p, sheet="S", header_row=1, pad_cols=["n1"], db=db, row_hash=True
    )

    with sqlite3.connect(str(db)) as conn:
        rows = conn.execute('SELECT row_hash, n1, n2 FROM "S" ORDER BY rowid').fetchall()

    # Storage should be TEXT and canonicalized
    assert rows[0][1] == "0"           # -0.0 -> "0"
    assert rows[0][2] == "1000000"     # no scientific notation
    assert rows[1][1] == "0"           # filled down as "0"

    # Hashes of ["0","1000000"] and ["0", "1000000"] should match our own recomputation
    expected = sha256_hex(canon_list(["0", "1000000"]))
    assert rows[0][0] == expected
    assert rows[1][0] == expected

def _write_wb_simple(tmp_path: Path, title="S"):
    p = tmp_path / "in_cli.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    ws.append(["A", "B"])
    ws.append(["x1", "y1"])
    ws.append([None, "y2"])
    ws.append(["x3", None])
    wb.save(p)
    return p


def test_cli_db_happy_path_prints_summary(tmp_path, capsys):
    """End-to-end CLI: db subcommand prints a success line and creates DB."""
    infile = _write_wb_simple(tmp_path, title="Sheet1")
    db = tmp_path / "out.db"
    cmd = [
        sys.executable, "-m", "xlfilldown.cli", "db",
        "--infile", str(infile),
        "--insheet", "Sheet1",
        "--header-row", "1",
        "--fill-cols", '["A"]',
        "--db", str(db),
        "--table", "t",
        "--row-hash",
        "--excel-row-numbers",
        "--if-exists", "replace",
        "--batch-size", "2",
    ]
    # Use subprocess to mimic real CLI
    r = subprocess.run(cmd, capture_output=True, text=True)
    assert r.returncode == 0, r.stderr
    assert "Loaded 'Sheet1' →" in r.stdout
    assert "(cols=2 (+ excel_row, row_hash);" in r.stdout
    assert db.exists()
    # sanity: db has rows
    with sqlite3.connect(str(db)) as conn:
        n = conn.execute("SELECT COUNT(*) FROM t;").fetchone()[0]
    assert n == 3


def test_cli_xlsx_happy_path_prints_summary(tmp_path):
    """End-to-end CLI: xlsx subcommand prints a success line and writes workbook."""
    infile = _write_wb_simple(tmp_path, title="S1")
    outfile = tmp_path / "out.xlsx"
    cmd = [
        sys.executable, "-m", "xlfilldown.cli", "xlsx",
        "--infile", str(infile),
        "--insheet", "S1",
        "--header-row", "1",
        "--fill-cols", '["A"]',
        "--outfile", str(outfile),
        "--outsheet", "Processed",
        "--row-hash",
        "--excel-row-numbers",
        "--if-exists", "fail",
    ]
    r = subprocess.run(cmd, capture_output=True, text=True)
    assert r.returncode == 0, r.stderr
    assert "Wrote 'S1' →" in r.stdout
    assert "(cols=2 (+ excel_row, row_hash);" in r.stdout
    wb = openpyxl.load_workbook(outfile)
    assert "Processed" in wb.sheetnames
    ws = wb["Processed"]
    # header + 3 data rows
    assert ws.max_row == 4
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert header == ["row_hash", "excel_row", "A", "B"]


def test_cli_pad_cols_bad_json_exits(tmp_path):
    """CLI should exit with a clear message if --fill-cols JSON is invalid."""
    infile = _write_wb_simple(tmp_path, title="S")
    db = tmp_path / "o.db"
    cmd = [
        sys.executable, "-m", "xlfilldown.cli", "db",
        "--infile", str(infile),
        "--insheet", "S",
        "--header-row", "1",
        "--fill-cols", 'not-a-json',
        "--db", str(db),
    ]
    r = subprocess.run(cmd, capture_output=True, text=True)
    assert r.returncode != 0
    assert "must be a valid JSON list of header names" in (r.stderr + r.stdout)


def test_cli_pad_cols_empty_list_exits(tmp_path):
    """CLI should reject an empty list for --fill-cols."""
    infile = _write_wb_simple(tmp_path, title="S")
    db = tmp_path / "o.db"
    cmd = [
        sys.executable, "-m", "xlfilldown.cli", "db",
        "--infile", str(infile),
        "--insheet", "S",
        "--header-row", "1",
        "--fill-cols", "[]",
        "--db", str(db),
    ]
    r = subprocess.run(cmd, capture_output=True, text=True)
    assert r.returncode != 0
    assert "--fill-cols cannot be empty" in (r.stderr + r.stdout)


def test_sqlite_append_exact_schema_match_succeeds(tmp_path):
    """Append works only when schema matches exactly (including optional columns and order)."""
    infile = _make_workbook(tmp_path)
    db = tmp_path / "samedb.db"

    # First load with row_hash + excel_row
    s1 = ingest_excel_to_sqlite(
        file=infile, sheet="Sheet1", header_row=1, pad_cols=["tier0"],
        db=db, table="t", row_hash=True, excel_row_numbers=True,
        if_exists="replace"
    )
    assert s1["rows_ingested"] == 6

    # Append with identical options → should succeed and double rows
    s2 = ingest_excel_to_sqlite(
        file=infile, sheet="Sheet1", header_row=1, pad_cols=["tier0"],
        db=db, table="t", row_hash=True, excel_row_numbers=True,
        if_exists="append"
    )
    assert s2["rows_ingested"] == 6
    with sqlite3.connect(str(db)) as conn:
        n = conn.execute("SELECT COUNT(*) FROM t;").fetchone()[0]
    assert n == 12


def test_sqlite_batch_size_one(tmp_path):
    """Batch size of 1 should still ingest all rows correctly."""
    infile = _make_workbook(tmp_path)
    db = tmp_path / "b1.db"
    s = ingest_excel_to_sqlite(
        file=infile, sheet="Sheet1", header_row=1, pad_cols=["tier0"],
        db=db, table="t", batch_size=1, if_exists="replace"
    )
    assert s["rows_ingested"] == 6
    with sqlite3.connect(str(db)) as conn:
        n = conn.execute("SELECT COUNT(*) FROM t;").fetchone()[0]
    assert n == 6


def test_excel_if_exists_replace_recreates(tmp_path):
    """'replace' at the sheet level should recreate the sheet cleanly (no residual rows)."""
    infile = _make_workbook(tmp_path)
    out = tmp_path / "r.xlsx"

    # first write
    ingest_excel_to_excel(
        file=infile, sheet="Sheet1", header_row=1, pad_cols=["tier0"],
        outfile=out, outsheet="Out", if_exists="fail"
    )
    # second write with replace
    ingest_excel_to_excel(
        file=infile, sheet="Sheet1", header_row=1, pad_cols=["tier0"],
        outfile=out, outsheet="Out", if_exists="replace"
    )

    wb = openpyxl.load_workbook(out)
    ws = wb["Out"]
    # header + 6 rows, not 12
    assert ws.max_row == 7


def test_pad_cols_duplicates_are_deduped_preserving_first(tmp_path):
    """Duplicate names in --fill-cols should be de-duped while preserving order of first occurrence."""
    p = tmp_path / "in_dupcols.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["A", "B", "C"])
    ws.append(["g1", None, "v1"])
    ws.append([None, "b1", "v2"])
    wb.save(p)

    db = tmp_path / "o.db"
    summary = ingest_excel_to_sqlite(
        file=p, sheet="S", header_row=1,
        pad_cols=["A", "B", "A", "B"],    # duplicates
        db=db, table="t", if_exists="replace"
    )
    assert summary["rows_ingested"] == 2
    with sqlite3.connect(str(db)) as conn:
        rows = conn.execute('SELECT A, B, C FROM t ORDER BY rowid').fetchall()
    # A should fill down to 'g1', B should carry b1 only on second row
    assert rows == [("g1", None, "v1"), ("g1", "b1", "v2")]


def test_require_non_null_after_padding_keeps_row(tmp_path):
    """Rows that become non-null after fill-down should pass require_non_null."""
    p = tmp_path / "in_req.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["Grp", "Val"])
    ws.append(["g1", None])   # header-ish -> will carry 'g1'
    ws.append([None, "v1"])   # after padding Grp='g1', so row is kept
    wb.save(p)

    db = tmp_path / "o.db"
    s = ingest_excel_to_sqlite(
        file=p, sheet="S", header_row=1, pad_cols=["Grp"],
        db=db, table="t", require_non_null=["Grp"], if_exists="replace"
    )
    assert s["rows_ingested"] == 2
    with sqlite3.connect(str(db)) as conn:
        rows = conn.execute('SELECT "Grp","Val" FROM t ORDER BY rowid').fetchall()
    assert rows == [("g1", None), ("g1", "v1")]


def test_sheet_not_found_raises(tmp_path):
    """Graceful error when the sheet name is wrong."""
    infile = _write_wb_simple(tmp_path, title="Good")
    with pytest.raises(ValueError, match="Sheet not found"):
        ingest_excel_to_sqlite(
            file=infile, sheet="BAD", header_row=1, pad_cols=["A"], db=tmp_path / "o.db"
        )


def test_all_blank_header_row_rejected(tmp_path):
    """A header row that normalizes to all blanks should raise."""
    p = tmp_path / "in_blankhdr.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["  ", None, "nan", "   "])  # all blank/nan → normalize to empty
    ws.append(["x", "y", "z", "w"])
    wb.save(p)

    with pytest.raises(ValueError, match="No non-empty headers"):
        ingest_excel_to_sqlite(
            file=p, sheet="S", header_row=1, pad_cols=["x"], db=tmp_path / "o.db"
        )


def test_cli_version_flag(tmp_path):
    """`xlfilldown --version` should print a version and exit 0."""
    r = subprocess.run([sys.executable, "-m", "xlfilldown.cli", "--version"],
                       capture_output=True, text=True)
    assert r.returncode == 0
    assert r.stdout.strip()  # non-empty


def test_cli_no_subcommand_shows_help(tmp_path):
    """Running cli without a subcommand should print help and exit 0."""
    r = subprocess.run([sys.executable, "-m", "xlfilldown.cli"],
                       capture_output=True, text=True)
    assert r.returncode == 0
    assert "db" in r.stdout and "xlsx" in r.stdout

def _mk_in(tmp_path: Path):
    p = tmp_path / "in.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["A"])
    ws.append(["x"])
    wb.save(p)
    wb.close()
    return p

def test_dest_workbook_is_unlocked_after_write(tmp_path):
    src = _mk_in(tmp_path)
    out = tmp_path / "out.xlsx"

    ingest_excel_to_excel(
        file=src, sheet="S", header_row=1, pad_cols=["A"],
        outfile=out, outsheet="Processed", if_exists="replace"
    )

    # If the destination workbook isn't closed, Windows would fail to rename.
    moved = tmp_path / "out_moved.xlsx"
    shutil.move(out, moved)  # should not raise

    # And we can write a new file at the original path again
    ingest_excel_to_excel(
        file=src, sheet="S", header_row=1, pad_cols=["A"],
        outfile=out, outsheet="Processed", if_exists="replace"
    )

def test_source_workbook_is_unlocked_after_read(tmp_path):
    src = _mk_in(tmp_path)
    out = tmp_path / "out.xlsx"

    ingest_excel_to_excel(
        file=src, sheet="S", header_row=1, pad_cols=["A"],
        outfile=out, outsheet="Processed", if_exists="replace"
    )

    # Being able to overwrite the source implies it’s closed.
    wb = openpyxl.Workbook()
    wb.save(src)  # should not raise
    wb.close()


def test_pad_cols_letters_rejects_empty_header(tmp_path, capsys):
    import openpyxl, subprocess, sys
    p = tmp_path / "in.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["A", "   ", "C"])   # B-header is whitespace → empty after normalization
    ws.append(["x", "y", "z"])
    wb.save(p); wb.close()

    cmd = [
        sys.executable, "-m", "xlfilldown.cli", "db",
        "--infile", str(p), "--insheet", "S", "--header-row", "1",
        "--fill-cols-letters", "B",            # points to empty header
        "--db", str(tmp_path / "o.db"),
    ]
    r = subprocess.run(cmd, capture_output=True, text=True)
    assert r.returncode != 0
    assert "refers to an empty header cell" in (r.stdout + r.stderr)

def test_cli_pad_cols_letters_happy_path_db(tmp_path):
    """Letters (A/C/…) resolve to headers on --header-row and padding works."""
    # Build a small workbook with 3 columns so we can reference A and C
    p = tmp_path / "letters_ok.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["Grp", "Val", "Note"])     # header row
    ws.append(["g1", None, "n1"])         # row 2: Grp present
    ws.append([None, "v2", "n2"])         # row 3: Grp should fill down to g1
    wb.save(p); wb.close()

    db = tmp_path / "out.db"
    cmd = [
        sys.executable, "-m", "xlfilldown.cli", "db",
        "--infile", str(p),
        "--insheet", "S",
        "--header-row", "1",
        "--fill-cols-letters", "A",   # resolves to "Grp"
        "--db", str(db),
        "--table", "t",
        "--if-exists", "replace",
    ]
    r = subprocess.run(cmd, capture_output=True, text=True)
    assert r.returncode == 0, r.stderr

    # Verify padding (Grp carried to second row). Also verifies the letters→headers mapping worked.
    with sqlite3.connect(str(db)) as conn:
        rows = conn.execute('SELECT "Grp","Val","Note" FROM t ORDER BY rowid').fetchall()
    assert rows == [("g1", None, "n1"), ("g1", "v2", "n2")]


def test_cli_pad_cols_letters_out_of_range_errors(tmp_path):
    """Letter past the header width should error with an 'out of range' message."""
    p = tmp_path / "letters_oor.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["A", "B", "C"])   # only 3 columns in header row
    ws.append([1, 2, 3])
    wb.save(p); wb.close()

    db = tmp_path / "o.db"
    cmd = [
        sys.executable, "-m", "xlfilldown.cli", "db",
        "--infile", str(p),
        "--insheet", "S",
        "--header-row", "1",
        "--fill-cols-letters", "ZZ",   # way out of range
        "--db", str(db),
    ]
    r = subprocess.run(cmd, capture_output=True, text=True)
    assert r.returncode != 0
    msg = (r.stderr + r.stdout)
    assert "out of range" in msg and "header row" in msg


def test_cli_pad_cols_letters_mutually_exclusive_errors(tmp_path):
    """Providing both --fill-cols and --fill-cols-letters must fail clearly."""
    p = tmp_path / "letters_me.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["Grp", "Val"])
    ws.append(["g1", "v1"])
    wb.save(p); wb.close()

    db = tmp_path / "o.db"
    cmd = [
        sys.executable, "-m", "xlfilldown.cli", "db",
        "--infile", str(p),
        "--insheet", "S",
        "--header-row", "1",
        "--fill-cols", '["Grp"]',
        "--fill-cols-letters", "A",
        "--db", str(db),
    ]
    r = subprocess.run(cmd, capture_output=True, text=True)
    assert r.returncode != 0
    assert "Use only one of --fill-cols or --fill-cols-letters" in (r.stderr + r.stdout)

# --- New tests for header-row bounds, CLI guard, and write_only sheet handling ---

def test_api_header_row_lt_one_errors(tmp_path):
    """API should raise a clear ValueError when header_row < 1."""
    p = tmp_path / "lt1.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["A"])
    ws.append(["x"])
    wb.save(p); wb.close()

    with pytest.raises(ValueError, match=r"Header row must be >= 1"):
        ingest_excel_to_sqlite(
            file=p, sheet="S", header_row=0, pad_cols=["A"], db=tmp_path / "o.db"
        )

def test_api_header_row_exceeds_max_errors_sqlite(tmp_path):
    """API should raise when header_row > max_row (via _sheet_headers)."""
    p = tmp_path / "hdr_oom.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["A"])        # max_row == 1
    ws.append(["x"])        # max_row == 2
    wb.save(p); wb.close()

    with pytest.raises(ValueError, match=r"exceeds sheet max row"):
        ingest_excel_to_sqlite(
            file=p, sheet="S", header_row=999, pad_cols=["A"], db=tmp_path / "o.db"
        )

def test_api_header_row_exceeds_max_errors_excel(tmp_path):
    """Same header_row > max_row behavior through Excel writer path."""
    p = tmp_path / "hdr_oom2.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["A"])
    ws.append(["x"])
    wb.save(p); wb.close()

    with pytest.raises(ValueError, match=r"exceeds sheet max row"):
        ingest_excel_to_excel(
            file=p, sheet="S", header_row=500, pad_cols=["A"],
            outfile=tmp_path / "out.xlsx", outsheet="Processed"
        )

def test_cli_header_row_zero_errors(tmp_path):
    """CLI should refuse --header-row 0 with a clear message."""
    infile = _mk_in(tmp_path)
    db = tmp_path / "o.db"
    cmd = [
        sys.executable, "-m", "xlfilldown.cli", "db",
        "--infile", str(infile),
        "--insheet", "S",
        "--header-row", "0",
        "--fill-cols", '["A"]',
        "--db", str(db),
    ]
    r = subprocess.run(cmd, capture_output=True, text=True)
    assert r.returncode != 0
    assert "--header-row must be >= 1" in (r.stderr + r.stdout)

def test_cli_letters_out_of_range_mentions_header_row_and_headered_cols(tmp_path):
    """Out-of-range letter error should mention header row context and headered columns."""
    p = tmp_path / "letters_oor2.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    # Only 2 headers; later data shouldn't matter for mapping by header row
    ws.append(["H1", "H2"])
    ws.append([1, 2])
    wb.save(p); wb.close()

    cmd = [
        sys.executable, "-m", "xlfilldown.cli", "db",
        "--infile", str(p),
        "--insheet", "S",
        "--header-row", "1",
        "--fill-cols-letters", "ZZ",
        "--db", str(tmp_path / "o.db"),
    ]
    r = subprocess.run(cmd, capture_output=True, text=True)
    msg = (r.stderr + r.stdout)
    assert r.returncode != 0
    assert "out of range" in msg and "header row 1" in msg
    # new friendlier hint
    assert "only headered columns are ingested" in msg

def test_new_workbook_has_no_extra_default_sheet(tmp_path):
    """When creating a new workbook in write_only mode, only the target sheet should exist."""
    src = _mk_in(tmp_path)  # creates an input with sheet 'S'
    out = tmp_path / "fresh.xlsx"

    # Write a new output workbook/sheet
    ingest_excel_to_excel(
        file=src, sheet="S", header_row=1, pad_cols=["A"],
        outfile=out, outsheet="Processed", if_exists="fail"
    )

    wb = openpyxl.load_workbook(out)
    try:
        # Should have exactly one sheet named 'Processed'
        assert wb.sheetnames == ["Processed"]
        ws = wb["Processed"]
        # Header must be present
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        assert header == ["A"]  # no row_hash/excel_row here
    finally:
        wb.close()

def test_cli_letters_header_row_exceeds_max_errors(tmp_path):
    """Letters path should error if header_row is beyond max_row (CLI)."""
    p = tmp_path / "hdr_oor_letters.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["A"])
    wb.save(p); wb.close()

    cmd = [
        sys.executable, "-m", "xlfilldown.cli", "db",
        "--infile", str(p),
        "--insheet", "S",
        "--header-row", "999",
        "--fill-cols-letters", "A",
        "--db", str(tmp_path / "o.db"),
    ]
    r = subprocess.run(cmd, capture_output=True, text=True)
    assert r.returncode != 0
    assert "exceeds sheet 'S' max row" in (r.stderr + r.stdout)


def test_excel_multiple_sheets_in_same_workbook(tmp_path):
    """Two invocations writing to the same outfile with different outsheet names should coexist."""
    infile = _make_workbook(tmp_path)
    outfile = tmp_path / "multi.xlsx"

    # First write to sheet "S1"
    ingest_excel_to_excel(
        file=infile,
        sheet="Sheet1",
        header_row=1,
        pad_cols=["tier0"],
        outfile=outfile,
        outsheet="S1",
        if_exists="fail",
    )

    # Second write to sheet "S2" in the same workbook
    ingest_excel_to_excel(
        file=infile,
        sheet="Sheet1",
        header_row=1,
        pad_cols=["tier0"],
        outfile=outfile,
        outsheet="S2",
        if_exists="fail",
    )

    # Validate workbook contains both sheets with the correct row counts
    wb = openpyxl.load_workbook(outfile)
    try:
        assert set(wb.sheetnames) == {"S1", "S2"}
        for name in ("S1", "S2"):
            ws = wb[name]
            # 1 header + 6 data rows
            assert ws.max_row == 7
            header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            assert header == ["tier0", "tier1"]
    finally:
        wb.close()

def test_cli_fill_cols_happy_path_db(tmp_path):
    p = tmp_path / "in.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "S"
    ws.append(["A","B"])
    ws.append(["x1","y1"])
    ws.append([None,"y2"])      # A should fill to x1
    wb.save(p); wb.close()

    db = tmp_path / "o.db"
    r = subprocess.run([
        sys.executable, "-m", "xlfilldown.cli", "db",
        "--infile", str(p), "--insheet", "S", "--header-row", "1",
        "--fill-cols", '["A"]',
        "--db", str(db), "--table", "t", "--if-exists", "replace",
    ], capture_output=True, text=True)
    assert r.returncode == 0, r.stderr
    with sqlite3.connect(str(db)) as conn:
        rows = conn.execute('SELECT "A","B" FROM t ORDER BY rowid').fetchall()
    assert rows == [("x1","y1"), ("x1","y2")]


def test_cli_fill_cols_letters_happy_path_db(tmp_path):
    p = tmp_path / "in2.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "S"
    ws.append(["Grp","Val","Note"])
    ws.append(["g1", None, "n1"])
    ws.append([None,"v2","n2"])
    wb.save(p); wb.close()

    db = tmp_path / "o2.db"
    r = subprocess.run([
        sys.executable, "-m", "xlfilldown.cli", "db",
        "--infile", str(p), "--insheet", "S", "--header-row", "1",
        "--fill-cols-letters", "A",     # => "Grp"
        "--db", str(db), "--table", "t", "--if-exists", "replace",
    ], capture_output=True, text=True)
    assert r.returncode == 0, r.stderr
    with sqlite3.connect(str(db)) as conn:
        rows = conn.execute('SELECT "Grp","Val","Note" FROM t ORDER BY rowid').fetchall()
    assert rows == [("g1", None, "n1"), ("g1","v2","n2")]



def test_require_non_null_letters_merges_with_names_cli(tmp_path):
    p = tmp_path / "in_req_letters.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "S"
    ws.append(["Grp","A","B"])
    ws.append([None,"x",None])    # drop (Grp missing)
    ws.append(["g1",None,"y"])    # keep (Grp present & B present)
    wb.save(p); wb.close()

    db = tmp_path / "o_req.db"
    r = subprocess.run([
        sys.executable, "-m", "xlfilldown.cli", "db",
        "--infile", str(p), "--insheet", "S", "--header-row", "1",
        "--fill-cols", '["Grp"]',
        "--require-non-null", '["B"]',
        "--require-non-null-letters", "A",   # "Grp"
        "--db", str(db), "--table", "t", "--if-exists", "replace",
    ], capture_output=True, text=True)
    assert r.returncode == 0, r.stderr
    with sqlite3.connect(str(db)) as conn:
        rows = conn.execute('SELECT "Grp","A","B" FROM t ORDER BY rowid').fetchall()
    assert rows == [("g1", None, "y")]


def test_cli_fill_cols_and_letters_mutually_exclusive(tmp_path):
    p = tmp_path / "in_me.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "S"
    ws.append(["A"]); ws.append(["x"])
    wb.save(p); wb.close()
    db = tmp_path / "o_me.db"
    r = subprocess.run([
        sys.executable, "-m", "xlfilldown.cli", "db",
        "--infile", str(p), "--insheet", "S", "--header-row", "1",
        "--fill-cols", '["A"]',
        "--fill-cols-letters", "A",
        "--db", str(db),
    ], capture_output=True, text=True)
    assert r.returncode != 0
    assert "only one of --fill-cols or --fill-cols-letters" in (r.stdout + r.stderr)



def test_cli_fill_cols_empty_list_exits(tmp_path):
    p = tmp_path / "in_emptylist.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "S"
    ws.append(["A"]); ws.append(["x"])
    wb.save(p); wb.close()
    db = tmp_path / "o_empty.db"
    r = subprocess.run([
        sys.executable, "-m", "xlfilldown.cli", "db",
        "--infile", str(p), "--insheet", "S", "--header-row", "1",
        "--fill-cols", "[]",
        "--db", str(db),
    ], capture_output=True, text=True)
    assert r.returncode != 0
    assert "--fill-cols cannot be empty" in (r.stderr + r.stdout)



def test_fill_mode_hierarchical_order_matters_letters(tmp_path):
    p = tmp_path / "hier_order.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "S"
    ws.append(["Tier1","Tier2","Val"])
    ws.append(["A", None, "v1"])     # new Tier1
    ws.append([None,"x","v2"])       # detail under A: Tier2=x
    ws.append(["B", None, "v3"])     # new Tier1 -> resets Tier2 carry
    ws.append([None,None,"v4"])      # this row exposes the reset behavior
    wb.save(p); wb.close()

    # Order: Tier1 > Tier2 (A then B resets Tier2)
    db1 = tmp_path / "o_hier1.db"
    subprocess.run([
        sys.executable, "-m", "xlfilldown.cli", "db",
        "--infile", str(p), "--insheet", "S", "--header-row", "1",
        "--fill-cols-letters", "A", "B",
        "--db", str(db1), "--table", "t", "--if-exists", "replace",
    ], check=True, capture_output=True, text=True)
    with sqlite3.connect(str(db1)) as conn:
        rows1 = conn.execute('SELECT "Tier1","Tier2","Val" FROM t ORDER BY rowid').fetchall()

    # Reverse order: Tier2 > Tier1 (no reset of Tier2 when Tier1 changes)
    db2 = tmp_path / "o_hier2.db"
    subprocess.run([
        sys.executable, "-m", "xlfilldown.cli", "db",
        "--infile", str(p), "--insheet", "S", "--header-row", "1",
        "--fill-cols-letters", "B", "A",
        "--db", str(db2), "--table", "t", "--if-exists", "replace",
    ], check=True, capture_output=True, text=True)
    with sqlite3.connect(str(db2)) as conn:
        rows2 = conn.execute('SELECT "Tier1","Tier2","Val" FROM t ORDER BY rowid').fetchall()

    # On the fourth data row (v4), hierarchical order changes the Tier2 value
    # rows are: v1, v2, v3, v4
    assert rows1[3] == ("B", None, "v4")     # Tier2 reset under Tier1 change
    assert rows2[3] == ("B", "x", "v4")      # Tier2 carried because Tier2 is higher tier


def test_fill_mode_independent_order_does_not_matter(tmp_path):
    p = tmp_path / "indep_order.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "S"
    ws.append(["T1","T2","Val"])
    ws.append(["A", None, "v1"])
    ws.append([None,"x","v2"])
    ws.append(["B", None, "v3"])
    ws.append([None,None,"v4"])
    wb.save(p); wb.close()

    def run_with_order(order, dbpath):
        r = subprocess.run([
            sys.executable, "-m", "xlfilldown.cli", "db",
            "--infile", str(p), "--insheet", "S", "--header-row", "1",
            "--fill-cols-letters", *order,
            "--fill-mode", "independent",
            "--db", str(dbpath), "--table", "t", "--if-exists", "replace",
        ], capture_output=True, text=True)
        assert r.returncode == 0, r.stderr
        with sqlite3.connect(str(dbpath)) as conn:
            return conn.execute('SELECT "T1","T2","Val" FROM t ORDER BY rowid').fetchall()

    rows1 = run_with_order(["A","B"], tmp_path / "o_ind1.db")
    rows2 = run_with_order(["B","A"], tmp_path / "o_ind2.db")
    assert rows1 == rows2



def test_require_non_null_letters_empty_header_errors(tmp_path):
    p = tmp_path / "req_empty_hdr.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "S"
    ws.append(["A", "   ", "C"])   # B header is whitespace -> empty after normalization
    ws.append(["x", "y", "z"])
    wb.save(p); wb.close()

    db = tmp_path / "o_req_empty_hdr.db"
    r = subprocess.run([
        sys.executable, "-m", "xlfilldown.cli", "db",
        "--infile", str(p), "--insheet", "S", "--header-row", "1",
        "--fill-cols", '["A"]',
        "--require-non-null-letters", "B",
        "--db", str(db),
    ], capture_output=True, text=True)
    assert r.returncode != 0
    assert "refers to an empty header cell" in (r.stdout + r.stderr)



def test_fill_mode_default_is_hierarchical(tmp_path):
    p = tmp_path / "default_mode.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "S"
    ws.append(["T1","T2"])
    ws.append(["A", None])   # row 2
    ws.append([None, "x"])   # row 3 -> carries T1 = "A", sets T2 = "x"
    ws.append(["B", None])   # row 4 -> T1 changes to "B" and resets lower tier T2 to None
    ws.append([None, None])  # row 5 -> completely empty spacer row (no fill applied)
    wb.save(p); wb.close()

    db = tmp_path / "o_default.db"
    r = subprocess.run([
        sys.executable, "-m", "xlfilldown.cli", "db",
        "--infile", str(p), "--insheet", "S", "--header-row", "1",
        "--fill-cols", '["T1","T2"]',
        "--db", str(db), "--table", "t", "--if-exists", "replace",
    ], capture_output=True, text=True)
    assert r.returncode == 0, r.stderr
    with sqlite3.connect(str(db)) as conn:
        rows = conn.execute('SELECT "T1","T2" FROM t ORDER BY rowid').fetchall()

    # rows are:
    # 0: ("A", None)
    # 1: ("A", "x")
    # 2: ("B", None)      <- hierarchical reset observed here
    # 3: (None, None)     <- spacer row preserved with no fill
    assert rows[2] == ("B", None)
    assert rows[3] == (None, None)





def test_cli_require_non_null_bad_json_exits(tmp_path):
    p = tmp_path / "in_badjson.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "S"
    ws.append(["A"]); ws.append(["x"])
    wb.save(p); wb.close()
    db = tmp_path / "o_badjson.db"
    r = subprocess.run([
        sys.executable, "-m", "xlfilldown.cli", "db",
        "--infile", str(p), "--insheet", "S", "--header-row", "1",
        "--fill-cols", '["A"]',
        "--require-non-null", 'not-json',
        "--db", str(db),
    ], capture_output=True, text=True)
    assert r.returncode != 0
    assert "--require-non-null must be a valid JSON list of header names" in (r.stderr + r.stdout)



def test_require_non_null_merge_order_and_dedup(tmp_path):
    p = tmp_path / "merge_req.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "S"
    ws.append(["A","B","C"])
    ws.append(["g", None, "x"])
    ws.append(["g", "y",  None])
    wb.save(p); wb.close()

    db = tmp_path / "m.db"
    r = subprocess.run([
        sys.executable, "-m", "xlfilldown.cli", "db",
        "--infile", str(p), "--insheet", "S", "--header-row", "1",
        "--fill-cols", '["A","B","C"]',
        "--require-non-null", '["C","A","C"]',    # duplicates + specific order
        "--require-non-null-letters", "B",        # merged at the end if new
        "--db", str(db), "--table", "t", "--if-exists", "replace",
    ], capture_output=True, text=True)
    assert r.returncode == 0, r.stderr
    # Enforce semantics (we don't read the merged list directly; behavior shows it)
    # Row1: A=g, B=None, C=x -> B missing -> DROPPED
    # Row2: A=g, B=y,   C=None -> C missing -> DROPPED
    with sqlite3.connect(str(db)) as conn:
        n = conn.execute('SELECT COUNT(*) FROM t').fetchone()[0]
    assert n == 0



def test_hierarchical_three_tiers_mid_level_reset(tmp_path):
    p = tmp_path / "3tiers.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "S"
    ws.append(["T1","T2","T3","Val"])
    ws.append(["A", None, None, "v1"])
    ws.append([None,"x",  "p",  "v2"])  # set T2=x, T3=p under T1=A
    ws.append([None,"y",  None, "v3"])  # change T2 to y -> reset T3 only
    wb.save(p); wb.close()

    db = tmp_path / "3.db"
    subprocess.run([
        sys.executable, "-m", "xlfilldown.cli", "db",
        "--infile", str(p), "--insheet", "S", "--header-row", "1",
        "--fill-cols", '["T1","T2","T3"]',
        "--db", str(db), "--table", "t", "--if-exists", "replace",
    ], check=True, capture_output=True, text=True)
    with sqlite3.connect(str(db)) as conn:
        rows = conn.execute('SELECT "T1","T2","T3","Val" FROM t ORDER BY rowid').fetchall()
    assert rows == [
        ("A", None, None, "v1"),
        ("A", "x",  "p",  "v2"),
        ("A", "y",  None, "v3"),  # T1 carried, T2 changed, T3 reset
    ]




def test_independent_mode_ignores_fill_cols_order(tmp_path):
    p = tmp_path / "indep_order2.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "S"
    ws.append(["A","B","V"])
    ws.append(["g1", None, "v1"])
    ws.append([None,"x",  "v2"])
    ws.append([None,None,"v3"])
    wb.save(p); wb.close()

    def run(fill_cols, dbpath):
        subprocess.run([
            sys.executable, "-m", "xlfilldown.cli", "db",
            "--infile", str(p), "--insheet", "S", "--header-row", "1",
            "--fill-cols", fill_cols,
            "--fill-mode", "independent",
            "--db", str(dbpath), "--table", "t", "--if-exists", "replace",
        ], check=True, capture_output=True, text=True)
        with sqlite3.connect(str(dbpath)) as conn:
            return conn.execute('SELECT "A","B","V" FROM t ORDER BY rowid').fetchall()

    rows1 = run('["A","B"]', tmp_path / "i1.db")
    rows2 = run('["B","A"]', tmp_path / "i2.db")
    assert rows1 == rows2


def test_date_canonicalization_and_hash(tmp_path):
    p = tmp_path / "dates.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "S"
    ws.append(["Grp","When","Txt"])
    # openpyxl stores python date/datetime objects; we use isoformat
    d1 = datetime(2024, 5, 17, 13, 45, 0)
    ws.append(["g1", d1, "a"])
    ws.append([None, None, "b"])  # When should fill
    wb.save(p); wb.close()

    db = tmp_path / "dates.db"
    summary = ingest_excel_to_sqlite(
        file=p, sheet="S", header_row=1, pad_cols=["Grp","When"],
        db=db, table="t", row_hash=True, if_exists="replace"
    )
    assert summary["rows_ingested"] == 2
    with sqlite3.connect(str(db)) as conn:
        rows = conn.execute('SELECT row_hash, "Grp","When","Txt" FROM t ORDER BY rowid').fetchall()

    # Expect isoformat for When, identical across rows after fill
    assert rows[0][2] == d1.isoformat()
    assert rows[1][2] == d1.isoformat()

    # Hashes should differ only by Txt
    h1, h2 = rows[0][0], rows[1][0]
    assert h1 != h2

def test_fill_cols_letters_out_of_range_mentions_context(tmp_path):
    p = tmp_path / "oor.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "S"
    ws.append(["Only"])
    ws.append(["x"])
    wb.save(p); wb.close()

    db = tmp_path / "oor.db"
    r = subprocess.run([
        sys.executable, "-m", "xlfilldown.cli", "db",
        "--infile", str(p), "--insheet", "S", "--header-row", "1",
        "--fill-cols-letters", "ZZ",
        "--db", str(db),
    ], capture_output=True, text=True)
    assert r.returncode != 0
    msg = (r.stderr + r.stdout)
    assert "out of range" in msg and "header row 1" in msg and "only headered columns are ingested" in msg


def test_fill_mode_hierarchical_excel_writer(tmp_path):
    p = tmp_path / "padmodes.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "S"
    ws.append(["Tier 1","Tier 2","Tier 3","Tier 4"])
    ws.append(["apple",  None,   None,   None])
    ws.append([None,     "red",  "sour", "value1"])
    ws.append(["potato", None,   None,   None])
    ws.append([None,     None,   None,   "value2"])
    ws.append([None,     "fried","yellow","value3"])
    wb.save(p); wb.close()

    out = tmp_path / "o.xlsx"
    s = ingest_excel_to_excel(
        file=p, sheet="S", header_row=1,
        pad_cols=["Tier 1","Tier 2","Tier 3"],
        outfile=out, outsheet="O",
        drop_blank_rows=True, require_non_null=["Tier 4"],
        pad_hierarchical=True, if_exists="replace"
    )
    wb2 = openpyxl.load_workbook(out)
    ws2 = wb2["O"]
    data = [[c.value for c in r] for r in ws2.iter_rows(min_row=2, max_row=1+s["rows_written"])]
    assert data == [
        ["apple","red","sour","value1"],
        ["potato", None, None, "value2"],
        ["potato","fried","yellow","value3"],
    ]
    wb2.close()


def test_drop_blank_rows_whitespace_with_fill_flags(tmp_path):
    p = tmp_path / "wsdrop2.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "S"
    ws.append(["grp","val"])
    ws.append(["   ", "x"])    # grp blank -> droppable when drop_blank_rows=True
    ws.append(["g1",  "y"])
    wb.save(p); wb.close()

    db = tmp_path / "wsdrop2.db"
    s = ingest_excel_to_sqlite(
        file=p, sheet="S", header_row=1, pad_cols=["grp"], db=db,
        drop_blank_rows=True, if_exists="replace"
    )
    with sqlite3.connect(str(db)) as conn:
        rows = conn.execute('SELECT grp, val FROM "S" ORDER BY rowid').fetchall()
    assert rows == [("g1","y")]














