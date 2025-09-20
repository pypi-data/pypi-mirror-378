import openpyxl

# Function to read the xlsx file and access specific sheets
def read_xlsx_and_process(file_path):
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)

    # Example of accessing specific sheets
    if 'ResourceDefinitions' in workbook.sheetnames:
        sheet = workbook['ResourceDefinitions']
        resource_definition_entities = process_sheet_resource_definitions(sheet)

    if 'ResourceLinks' in workbook.sheetnames:
        sheet = workbook['ResourceLinks']
        resource_link_entities = process_sheet_resource_links(sheet)

    if 'PatientData' in workbook.sheetnames:
        sheet = workbook['PatientData']
        patient_data_entities, num_entries = process_sheet_patient_data(sheet, resource_definition_entities)
    
    return {
        "resource_definition_entities": resource_definition_entities,
        "resource_link_entities": resource_link_entities,
        "patient_data_entities": patient_data_entities,
        "num_entries": num_entries
    }


# Function to process the specific sheet with 'Entity Name', 'ResourceType', and 'Profile(s)'
def process_sheet_resource_definitions(sheet):
    resource_definitions = []
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]  # Get headers

    for row in sheet.iter_rows(min_row=3, values_only=True):
        row_data = dict(zip(headers, row))  # Create a dictionary for each row
        if all(cell is None or cell == "" for cell in row_data.values()):
            continue
        # Split 'Profile(s)' column into a list of URLs
        if row_data.get("Profile(s)"):
            row_data["Profile(s)"] = [url.strip() for url in row_data["Profile(s)"].split(",")]

        resource_definitions.append(row_data)

    return resource_definitions

# Function to process the specific sheet with 'OriginResource', 'ReferencePath', and 'DestinationResource'
def process_sheet_resource_links(sheet):
    resource_links = []
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]  # Get headers
    for row in sheet.iter_rows(min_row=3, values_only=True):
        row_data = dict(zip(headers, row))  # Create a dictionary for each row
        if all(cell is None or cell == "" for cell in row_data):
            continue
        resource_links.append(row_data)

    return resource_links

# Function to process the "PatientData" sheet
def process_sheet_patient_data(sheet, resource_definition_entities):
    # Initialize the dictionary to store the processed data
    patient_data = {}
    # Extract the data from the first 6 rows (Entity To Query, JsonPath, etc.)
    for col in sheet.iter_cols(min_row=1, max_row=6, min_col=3, values_only=True):  # Start from 3rd column
        if all(entry is None for entry in col):
            continue
        entity_name = col[0]  # The entity name comes from the first row (Entity To Query)
        field_name = col[5]  #The "Data Element" comes from the fifth row
        if (entity_name is None or entity_name == "") and (field_name is not None and field_name != ""):
            print(f"WARNING: - Reading Patient Data Issue - {field_name} - 'Entity To Query' cell missing for column labelled '{field_name}', please provide entity name from the ResourceDefinitions tab.")

        if entity_name not in [entry['Entity Name'] for entry in resource_definition_entities]:
            print(f"WARNING: - Reading Patient Data Issue - {field_name} - 'Entity To Query' cell has entity named '{entity_name}', however, the ResourceDefinition tab has no matching resource. Please provide a corresponding entry in the ResourceDefinition tab.")
        # Create structure for this entity if not already present
        if entity_name not in patient_data:
            patient_data[entity_name] = {}

        # Add jsonpath, valuesets, and initialize an empty list for 'values'
        if field_name not in patient_data[entity_name]:
            patient_data[entity_name][field_name] = {
                "jsonpath": col[1],  # JsonPath from the second row
                "valueType": col[2], # Value Type from the third row
                "valuesets": col[3], # Value Set from the fourth row
                "values": []         # Initialize empty list for actual values
            }
    
    # Now process the rows starting from the 6th row (the actual data entries)
    num_entries = 0
    for row in sheet.iter_rows(min_row=7, values_only=True):  # Start from row 6 for actual data
        if all(cell is None for cell in row):
            continue
        num_entries = num_entries + 1
        entity_name = row[0]  # The entity name comes from the first column of each row
        for i, value in enumerate(row[2:], start=1):  # Iterate through the values in the columns
            entity_name = sheet.cell(row=1, column=i + 2).value
            field_name = sheet.cell(row=6, column=i + 2).value  # Get the Data Element for this column
            if entity_name in patient_data and field_name in patient_data[entity_name]:
                # Append the actual data values to the 'values' array
                patient_data[entity_name][field_name]["values"].append(value)
    return patient_data, num_entries