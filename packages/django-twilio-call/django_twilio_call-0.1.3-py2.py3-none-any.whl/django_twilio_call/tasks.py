"""Celery tasks for django-twilio-call package."""

import logging
import time
from datetime import datetime, timedelta
from typing import Dict, List, Optional

import requests
from celery import group, shared_task
from django.conf import settings
from django.core.cache import cache
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.utils import timezone

from .constants import DefaultValues, TimeIntervals

logger = logging.getLogger(__name__)


# ================================
# RECORDING PROCESSING TASKS
# ================================


@shared_task(bind=True, name="process_call_recording")
def process_call_recording(self, call_id: int, recording_data: Optional[Dict] = None):
    """Process and store call recording with transcription.

    Args:
        call_id: Call ID to process recording for
        recording_data: Optional recording data from webhook

    Returns:
        Dictionary with processing results

    """
    try:
        from .models import Call, CallRecording
        from .services.recording_service import recording_service

        call = Call.objects.get(id=call_id)

        # Update task progress
        self.update_state(state="PROGRESS", meta={"current": 1, "total": 5, "status": "Starting recording processing"})

        # If recording_data provided, create/update recording from webhook
        if recording_data:
            recording = recording_service.process_recording_callback(recording_data)
        else:
            # Find existing recording
            recording = CallRecording.objects.filter(call=call, status=CallRecording.Status.COMPLETED).first()

            if not recording:
                raise ValueError(f"No recording found for call {call.twilio_sid}")

        self.update_state(state="PROGRESS", meta={"current": 2, "total": 5, "status": "Recording retrieved"})

        # Download and store recording if needed
        if recording.url and not recording.file_size:
            try:
                response = requests.get(recording.url + ".mp3", timeout=DefaultValues.RECORDING_TIMEOUT)
                if response.status_code == 200:
                    recording.file_size = len(response.content)
                    recording.save(update_fields=["file_size"])
            except Exception as e:
                logger.warning(f"Failed to get file size for recording {recording.id}: {e}")

        self.update_state(state="PROGRESS", meta={"current": 3, "total": 5, "status": "Recording stored"})

        # Apply compliance and encryption
        recording_service._apply_compliance(recording)

        self.update_state(state="PROGRESS", meta={"current": 4, "total": 5, "status": "Compliance applied"})

        # Schedule transcription if enabled
        from django.conf import settings

        if getattr(settings, "ENABLE_TRANSCRIPTION", False):
            transcribe_recording.delay(recording.id)

        self.update_state(state="PROGRESS", meta={"current": 5, "total": 5, "status": "Transcription scheduled"})

        # Update call metadata
        call.metadata = call.metadata or {}
        call.metadata["recording_processed_at"] = timezone.now().isoformat()
        call.save(update_fields=["metadata"])

        logger.info(f"Successfully processed recording for call {call.twilio_sid}")

        return {
            "call_id": call_id,
            "recording_id": recording.id,
            "recording_duration": recording.duration,
            "file_size": recording.file_size,
            "processed_at": timezone.now().isoformat(),
        }

    except Call.DoesNotExist:
        logger.error(f"Call {call_id} not found")
        raise
    except Exception as e:
        logger.error(f"Failed to process recording for call {call_id}: {e}")
        raise


@shared_task(bind=True, name="transcribe_recording")
def transcribe_recording(self, recording_id: int, language: str = "en-US"):
    """Transcribe a call recording.

    Args:
        recording_id: Recording ID to transcribe
        language: Language for transcription

    Returns:
        Dictionary with transcription results

    """
    try:
        from .models import CallRecording
        from .services.recording_service import recording_service

        recording = CallRecording.objects.get(id=recording_id)

        # Update task progress
        self.update_state(state="PROGRESS", meta={"current": 1, "total": 3, "status": "Starting transcription"})

        # Check if already transcribed
        if recording.transcription:
            return {
                "recording_id": recording_id,
                "transcription": recording.transcription,
                "already_transcribed": True,
            }

        self.update_state(state="PROGRESS", meta={"current": 2, "total": 3, "status": "Processing audio"})

        # Perform transcription
        transcription = recording_service.transcribe_recording(recording_id, language)

        self.update_state(state="PROGRESS", meta={"current": 3, "total": 3, "status": "Transcription completed"})

        logger.info(f"Successfully transcribed recording {recording_id}")

        return {
            "recording_id": recording_id,
            "transcription": transcription,
            "language": language,
            "transcribed_at": timezone.now().isoformat(),
        }

    except CallRecording.DoesNotExist:
        logger.error(f"Recording {recording_id} not found")
        raise
    except Exception as e:
        logger.error(f"Failed to transcribe recording {recording_id}: {e}")
        raise


@shared_task(name="process_pending_recordings")
def process_pending_recordings():
    """Process all pending recordings that need processing."""
    from .models import CallRecording

    # Find recordings that are completed but not processed
    pending_recordings = CallRecording.objects.filter(
        status=CallRecording.Status.COMPLETED, transcription_status__in=["", "pending", None]
    ).exclude(call__metadata__has_key="recording_processed_at")[:10]  # Process max 10 at a time

    if not pending_recordings:
        return {"message": "No pending recordings to process"}

    # Create group of processing tasks
    job = group(process_call_recording.s(recording.call.id) for recording in pending_recordings)

    result = job.apply_async()

    return {
        "total_recordings": len(pending_recordings),
        "group_id": result.id,
        "started_at": timezone.now().isoformat(),
    }


@shared_task(name="transcribe_pending_recordings")
def transcribe_pending_recordings():
    """Transcribe all recordings that need transcription."""
    from .models import CallRecording

    # Find recordings that need transcription
    pending_transcriptions = CallRecording.objects.filter(
        status=CallRecording.Status.COMPLETED, transcription="", transcription_status__in=["", "pending", None]
    ).exclude(transcription_status="failed")[:5]  # Process max 5 at a time (transcription is expensive)

    if not pending_transcriptions:
        return {"message": "No pending transcriptions"}

    # Create group of transcription tasks
    job = group(transcribe_recording.s(recording.id) for recording in pending_transcriptions)

    result = job.apply_async()

    return {
        "total_transcriptions": len(pending_transcriptions),
        "group_id": result.id,
        "started_at": timezone.now().isoformat(),
    }


# ================================
# ANALYTICS AND REPORTING TASKS
# ================================


@shared_task(bind=True, name="generate_daily_report")
def generate_daily_report(self, date_str: Optional[str] = None):
    """Generate comprehensive daily analytics report.

    Args:
        date_str: Date string in YYYY-MM-DD format, defaults to yesterday

    Returns:
        Dictionary with report data and file paths

    """
    try:
        from .services.analytics_service import analytics_service

        # Parse date or use yesterday
        if date_str:
            report_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        else:
            report_date = (timezone.now() - timedelta(days=1)).date()

        start_datetime = timezone.make_aware(datetime.combine(report_date, datetime.min.time()))
        end_datetime = start_datetime + timedelta(days=1)

        self.update_state(state="PROGRESS", meta={"current": 1, "total": 6, "status": "Gathering call analytics"})

        # Generate call analytics
        call_analytics = analytics_service.get_call_analytics(start_date=start_datetime, end_date=end_datetime)

        self.update_state(state="PROGRESS", meta={"current": 2, "total": 6, "status": "Gathering agent analytics"})

        # Generate agent analytics
        agent_analytics = analytics_service.get_agent_analytics(start_date=start_datetime, end_date=end_datetime)

        self.update_state(state="PROGRESS", meta={"current": 3, "total": 6, "status": "Gathering queue analytics"})

        # Generate queue analytics
        queue_analytics = analytics_service.get_queue_analytics(start_date=start_datetime, end_date=end_datetime)

        self.update_state(state="PROGRESS", meta={"current": 4, "total": 6, "status": "Compiling report"})

        # Compile comprehensive report
        report_data = {
            "date": report_date.isoformat(),
            "generated_at": timezone.now().isoformat(),
            "call_analytics": call_analytics,
            "agent_analytics": agent_analytics,
            "queue_analytics": queue_analytics,
            "summary": {
                "total_calls": call_analytics["volume"]["total_calls"],
                "total_agents": agent_analytics["total_agents"],
                "total_queues": queue_analytics["total_queues"],
                "avg_service_level": call_analytics["queue"]["service_level"],
                "avg_abandonment_rate": call_analytics["performance"]["abandonment_rate"],
            },
        }

        self.update_state(state="PROGRESS", meta={"current": 5, "total": 6, "status": "Saving report"})

        # Save report to cache and optionally to file
        cache_key = f"daily_report_{report_date.isoformat()}"
        cache.set(cache_key, report_data, 86400 * 7)  # Cache for 7 days

        self.update_state(state="PROGRESS", meta={"current": 6, "total": 6, "status": "Report completed"})

        # Schedule email delivery if configured
        email_recipients = getattr(settings, "DAILY_REPORT_RECIPIENTS", [])
        if email_recipients:
            send_report_email.delay(report_data=report_data, recipients=email_recipients, report_type="daily")

        logger.info(f"Generated daily report for {report_date}")

        return {
            "date": report_date.isoformat(),
            "total_calls": report_data["summary"]["total_calls"],
            "cache_key": cache_key,
            "generated_at": report_data["generated_at"],
        }

    except Exception as e:
        logger.error(f"Failed to generate daily report: {e}")
        raise


@shared_task(bind=True, name="generate_weekly_report")
def generate_weekly_report(self, week_start_str: Optional[str] = None):
    """Generate weekly analytics report.

    Args:
        week_start_str: Week start date string, defaults to last week

    Returns:
        Dictionary with weekly report data

    """
    try:
        from .services.analytics_service import analytics_service

        # Calculate week start (Monday)
        if week_start_str:
            week_start = datetime.strptime(week_start_str, "%Y-%m-%d").date()
        else:
            today = timezone.now().date()
            days_since_monday = today.weekday()
            week_start = today - timedelta(days=days_since_monday + 7)  # Last week

        week_end = week_start + timedelta(days=7)
        start_datetime = timezone.make_aware(datetime.combine(week_start, datetime.min.time()))
        end_datetime = timezone.make_aware(datetime.combine(week_end, datetime.min.time()))

        self.update_state(state="PROGRESS", meta={"current": 1, "total": 4, "status": "Gathering weekly analytics"})

        # Generate analytics for the week
        call_analytics = analytics_service.get_call_analytics(start_date=start_datetime, end_date=end_datetime)

        agent_analytics = analytics_service.get_agent_analytics(start_date=start_datetime, end_date=end_datetime)

        self.update_state(state="PROGRESS", meta={"current": 2, "total": 4, "status": "Comparing with previous week"})

        # Compare with previous week
        prev_week_start = week_start - timedelta(days=7)
        prev_week_end = prev_week_start + timedelta(days=7)
        prev_start_datetime = timezone.make_aware(datetime.combine(prev_week_start, datetime.min.time()))
        prev_end_datetime = timezone.make_aware(datetime.combine(prev_week_end, datetime.min.time()))

        prev_call_analytics = analytics_service.get_call_analytics(
            start_date=prev_start_datetime, end_date=prev_end_datetime
        )

        self.update_state(state="PROGRESS", meta={"current": 3, "total": 4, "status": "Calculating trends"})

        # Calculate week-over-week changes
        wow_changes = {
            "calls": {
                "current": call_analytics["volume"]["total_calls"],
                "previous": prev_call_analytics["volume"]["total_calls"],
                "change_pct": _calculate_percentage_change(
                    prev_call_analytics["volume"]["total_calls"], call_analytics["volume"]["total_calls"]
                ),
            },
            "service_level": {
                "current": call_analytics["queue"]["service_level"],
                "previous": prev_call_analytics["queue"]["service_level"],
                "change_pct": _calculate_percentage_change(
                    prev_call_analytics["queue"]["service_level"], call_analytics["queue"]["service_level"]
                ),
            },
            "abandonment_rate": {
                "current": call_analytics["performance"]["abandonment_rate"],
                "previous": prev_call_analytics["performance"]["abandonment_rate"],
                "change_pct": _calculate_percentage_change(
                    prev_call_analytics["performance"]["abandonment_rate"],
                    call_analytics["performance"]["abandonment_rate"],
                ),
            },
        }

        report_data = {
            "week_start": week_start.isoformat(),
            "week_end": week_end.isoformat(),
            "generated_at": timezone.now().isoformat(),
            "call_analytics": call_analytics,
            "agent_analytics": agent_analytics,
            "week_over_week": wow_changes,
        }

        self.update_state(state="PROGRESS", meta={"current": 4, "total": 4, "status": "Weekly report completed"})

        # Cache the report
        cache_key = f"weekly_report_{week_start.isoformat()}"
        cache.set(cache_key, report_data, 86400 * 14)  # Cache for 14 days

        logger.info(f"Generated weekly report for week starting {week_start}")

        return {
            "week_start": week_start.isoformat(),
            "total_calls": call_analytics["volume"]["total_calls"],
            "cache_key": cache_key,
            "generated_at": report_data["generated_at"],
        }

    except Exception as e:
        logger.error(f"Failed to generate weekly report: {e}")
        raise


@shared_task(bind=True, name="generate_monthly_report")
def generate_monthly_report(self, year: Optional[int] = None, month: Optional[int] = None):
    """Generate monthly analytics report.

    Args:
        year: Year for report, defaults to last month
        month: Month for report, defaults to last month

    Returns:
        Dictionary with monthly report data

    """
    try:
        from .services.analytics_service import analytics_service

        # Calculate month
        if year and month:
            report_date = datetime(year, month, 1).date()
        else:
            today = timezone.now().date()
            if today.month == 1:
                report_date = datetime(today.year - 1, 12, 1).date()
            else:
                report_date = datetime(today.year, today.month - 1, 1).date()

        # Calculate month range
        if report_date.month == 12:
            next_month = datetime(report_date.year + 1, 1, 1).date()
        else:
            next_month = datetime(report_date.year, report_date.month + 1, 1).date()

        start_datetime = timezone.make_aware(datetime.combine(report_date, datetime.min.time()))
        end_datetime = timezone.make_aware(datetime.combine(next_month, datetime.min.time()))

        self.update_state(state="PROGRESS", meta={"current": 1, "total": 5, "status": "Gathering monthly analytics"})

        # Generate comprehensive monthly analytics
        call_analytics = analytics_service.get_call_analytics(start_date=start_datetime, end_date=end_datetime)

        agent_analytics = analytics_service.get_agent_analytics(start_date=start_datetime, end_date=end_datetime)

        queue_analytics = analytics_service.get_queue_analytics(start_date=start_datetime, end_date=end_datetime)

        self.update_state(state="PROGRESS", meta={"current": 2, "total": 5, "status": "Calculating daily breakdown"})

        # Calculate daily breakdown for the month
        daily_breakdown = []
        current_date = report_date
        while current_date < next_month.date():
            day_start = timezone.make_aware(datetime.combine(current_date, datetime.min.time()))
            day_end = day_start + timedelta(days=1)

            day_analytics = analytics_service.get_call_analytics(start_date=day_start, end_date=day_end)

            daily_breakdown.append(
                {
                    "date": current_date.isoformat(),
                    "total_calls": day_analytics["volume"]["total_calls"],
                    "completed_calls": day_analytics["volume"]["completed_calls"],
                    "avg_duration": day_analytics["duration"]["avg_duration"],
                    "service_level": day_analytics["queue"]["service_level"],
                }
            )

            current_date += timedelta(days=1)

        self.update_state(state="PROGRESS", meta={"current": 3, "total": 5, "status": "Comparing with previous month"})

        # Compare with previous month
        if report_date.month == 1:
            prev_month_date = datetime(report_date.year - 1, 12, 1).date()
            prev_next_month = datetime(report_date.year, 1, 1).date()
        else:
            prev_month_date = datetime(report_date.year, report_date.month - 1, 1).date()
            prev_next_month = report_date

        prev_start_datetime = timezone.make_aware(datetime.combine(prev_month_date, datetime.min.time()))
        prev_end_datetime = timezone.make_aware(datetime.combine(prev_next_month, datetime.min.time()))

        prev_call_analytics = analytics_service.get_call_analytics(
            start_date=prev_start_datetime, end_date=prev_end_datetime
        )

        self.update_state(state="PROGRESS", meta={"current": 4, "total": 5, "status": "Calculating trends"})

        # Calculate month-over-month changes
        mom_changes = {
            "calls": _calculate_percentage_change(
                prev_call_analytics["volume"]["total_calls"], call_analytics["volume"]["total_calls"]
            ),
            "service_level": _calculate_percentage_change(
                prev_call_analytics["queue"]["service_level"], call_analytics["queue"]["service_level"]
            ),
            "avg_duration": _calculate_percentage_change(
                prev_call_analytics["duration"]["avg_duration"], call_analytics["duration"]["avg_duration"]
            ),
        }

        report_data = {
            "year": report_date.year,
            "month": report_date.month,
            "month_name": report_date.strftime("%B"),
            "generated_at": timezone.now().isoformat(),
            "call_analytics": call_analytics,
            "agent_analytics": agent_analytics,
            "queue_analytics": queue_analytics,
            "daily_breakdown": daily_breakdown,
            "month_over_month": mom_changes,
        }

        self.update_state(state="PROGRESS", meta={"current": 5, "total": 5, "status": "Monthly report completed"})

        # Cache the report
        cache_key = f"monthly_report_{report_date.year}_{report_date.month:02d}"
        cache.set(cache_key, report_data, 86400 * 30)  # Cache for 30 days

        logger.info(f"Generated monthly report for {report_date.strftime('%B %Y')}")

        return {
            "year": report_date.year,
            "month": report_date.month,
            "total_calls": call_analytics["volume"]["total_calls"],
            "cache_key": cache_key,
            "generated_at": report_data["generated_at"],
        }

    except Exception as e:
        logger.error(f"Failed to generate monthly report: {e}")
        raise


@shared_task(bind=True, name="calculate_agent_metrics")
def calculate_agent_metrics(self, agent_id: int, date_range: Optional[Dict] = None):
    """Calculate comprehensive metrics for a specific agent.

    Args:
        agent_id: Agent ID to calculate metrics for
        date_range: Optional date range dict with 'start' and 'end' keys

    Returns:
        Dictionary with agent metrics

    """
    try:
        from .models import Agent
        from .services.analytics_service import analytics_service

        agent = Agent.objects.get(id=agent_id)

        # Parse date range or use today
        if date_range:
            start_date = datetime.fromisoformat(date_range["start"])
            end_date = datetime.fromisoformat(date_range["end"])
        else:
            end_date = timezone.now()
            start_date = end_date.replace(hour=0, minute=0, second=0, microsecond=0)

        self.update_state(
            state="PROGRESS",
            meta={"current": 1, "total": 4, "status": f"Calculating metrics for {agent.user.get_full_name()}"},
        )

        # Get agent analytics
        agent_analytics = analytics_service.get_agent_analytics(
            agent_id=agent_id, start_date=start_date, end_date=end_date
        )

        self.update_state(state="PROGRESS", meta={"current": 2, "total": 4, "status": "Calculating performance trends"})

        # Calculate 7-day trend
        week_ago = start_date - timedelta(days=7)
        trend_analytics = analytics_service.get_agent_analytics(
            agent_id=agent_id, start_date=week_ago, end_date=start_date
        )

        self.update_state(
            state="PROGRESS", meta={"current": 3, "total": 4, "status": "Compiling comprehensive metrics"}
        )

        # Compile comprehensive metrics
        if agent_analytics["agents"]:
            current_metrics = agent_analytics["agents"][0]
            trend_metrics = trend_analytics["agents"][0] if trend_analytics["agents"] else {}

            metrics = {
                "agent_id": agent_id,
                "agent_name": agent.user.get_full_name(),
                "period": {
                    "start": start_date.isoformat(),
                    "end": end_date.isoformat(),
                },
                "current_metrics": current_metrics,
                "trends": {
                    "calls_change": _calculate_percentage_change(
                        trend_metrics.get("calls", {}).get("total", 0), current_metrics["calls"]["total"]
                    ),
                    "aht_change": _calculate_percentage_change(
                        trend_metrics.get("calls", {}).get("avg_handling_time", 0),
                        current_metrics["calls"]["avg_handling_time"],
                    ),
                    "occupancy_change": _calculate_percentage_change(
                        trend_metrics.get("occupancy_rate", 0), current_metrics["occupancy_rate"]
                    ),
                },
                "calculated_at": timezone.now().isoformat(),
            }
        else:
            metrics = {
                "agent_id": agent_id,
                "agent_name": agent.user.get_full_name(),
                "error": "No data available for specified period",
            }

        self.update_state(state="PROGRESS", meta={"current": 4, "total": 4, "status": "Agent metrics completed"})

        # Cache metrics
        cache_key = f"agent_metrics_{agent_id}_{start_date.date()}_{end_date.date()}"
        cache.set(cache_key, metrics, 3600)  # Cache for 1 hour

        logger.info(f"Calculated metrics for agent {agent.user.get_full_name()}")

        return metrics

    except Agent.DoesNotExist:
        logger.error(f"Agent {agent_id} not found")
        raise
    except Exception as e:
        logger.error(f"Failed to calculate agent metrics: {e}")
        raise


@shared_task(name="calculate_hourly_metrics")
def calculate_hourly_metrics():
    """Calculate and cache hourly metrics for real-time dashboard."""
    try:
        from .services.analytics_service import analytics_service

        # Get metrics for the current hour
        now = timezone.now()
        hour_start = now.replace(minute=0, second=0, microsecond=0)
        hour_end = hour_start + timedelta(hours=1)

        # Get real-time metrics
        real_time_metrics = analytics_service.get_real_time_metrics()

        # Get hourly call analytics
        hourly_analytics = analytics_service.get_call_analytics(start_date=hour_start, end_date=hour_end)

        metrics = {
            "hour": hour_start.isoformat(),
            "real_time": real_time_metrics,
            "hourly": hourly_analytics,
            "calculated_at": now.isoformat(),
        }

        # Cache metrics for dashboard
        cache_key = f"hourly_metrics_{hour_start.strftime('%Y%m%d_%H')}"
        cache.set(cache_key, metrics, 3600)  # Cache for 1 hour

        return {
            "hour": hour_start.isoformat(),
            "total_calls_hour": hourly_analytics["volume"]["total_calls"],
            "cache_key": cache_key,
        }

    except Exception as e:
        logger.error(f"Failed to calculate hourly metrics: {e}")
        raise


# ================================
# DATA MAINTENANCE TASKS
# ================================


@shared_task(bind=True, name="cleanup_old_call_logs")
def cleanup_old_call_logs(self, days_to_keep: int = 90):
    """Clean up old call logs to manage database size.

    Args:
        days_to_keep: Number of days of logs to keep

    Returns:
        Dictionary with cleanup results

    """
    try:
        from .models import CallLog

        cutoff_date = timezone.now() - timedelta(days=days_to_keep)

        self.update_state(state="PROGRESS", meta={"current": 1, "total": 3, "status": "Identifying old call logs"})

        # Count logs to be deleted
        old_logs_count = CallLog.objects.filter(created_at__lt=cutoff_date).count()

        if old_logs_count == 0:
            return {"message": "No old call logs to clean up"}

        self.update_state(
            state="PROGRESS", meta={"current": 2, "total": 3, "status": f"Deleting {old_logs_count} old call logs"}
        )

        # Delete old logs in batches
        batch_size = 1000
        deleted_total = 0

        while True:
            batch_ids = list(
                CallLog.objects.filter(created_at__lt=cutoff_date).values_list("id", flat=True)[:batch_size]
            )

            if not batch_ids:
                break

            deleted_count, _ = CallLog.objects.filter(id__in=batch_ids).delete()
            deleted_total += deleted_count

            # Small delay to prevent overwhelming the database
            time.sleep(0.1)

        self.update_state(state="PROGRESS", meta={"current": 3, "total": 3, "status": "Cleanup completed"})

        logger.info(f"Cleaned up {deleted_total} call logs older than {days_to_keep} days")

        return {
            "deleted_count": deleted_total,
            "cutoff_date": cutoff_date.isoformat(),
            "days_kept": days_to_keep,
            "completed_at": timezone.now().isoformat(),
        }

    except Exception as e:
        logger.error(f"Failed to cleanup old call logs: {e}")
        raise


@shared_task(bind=True, name="archive_old_recordings")
def archive_old_recordings(self, days_to_keep: int = 365):
    """Archive old recordings based on retention policy.

    Args:
        days_to_keep: Number of days to keep recordings accessible

    Returns:
        Dictionary with archival results

    """
    try:
        from .models import CallRecording
        from .services.recording_service import recording_service

        cutoff_date = timezone.now() - timedelta(days=days_to_keep)

        self.update_state(
            state="PROGRESS", meta={"current": 1, "total": 4, "status": "Identifying recordings to archive"}
        )

        # Find recordings to archive
        recordings_to_archive = CallRecording.objects.filter(
            created_at__lt=cutoff_date, status=CallRecording.Status.COMPLETED
        ).exclude(status=CallRecording.Status.DELETED)

        total_recordings = recordings_to_archive.count()

        if total_recordings == 0:
            return {"message": "No recordings to archive"}

        self.update_state(
            state="PROGRESS", meta={"current": 2, "total": 4, "status": f"Archiving {total_recordings} recordings"}
        )

        archived_count = 0
        failed_count = 0

        # Process recordings in batches
        for recording in recordings_to_archive.iterator(chunk_size=100):
            try:
                # Check retention policy in metadata
                retention_date_str = recording.metadata.get("retention_date")
                if retention_date_str:
                    retention_date = datetime.fromisoformat(retention_date_str.replace("Z", "+00:00"))
                    if timezone.now() < retention_date:
                        continue  # Not yet time to archive

                # Archive the recording (move to cold storage or delete)
                if getattr(settings, "RECORDING_ARCHIVE_TO_COLD_STORAGE", False):
                    # Move to cold storage (implementation depends on storage backend)
                    recording.metadata["archived_at"] = timezone.now().isoformat()
                    recording.metadata["archive_location"] = "cold_storage"
                    recording.save(update_fields=["metadata"])
                else:
                    # Delete the recording
                    recording_service.delete_recording(
                        recording.id, reason=f"Automated archival after {days_to_keep} days"
                    )

                archived_count += 1

            except Exception as e:
                logger.error(f"Failed to archive recording {recording.id}: {e}")
                failed_count += 1

            # Update progress every 10 recordings
            if (archived_count + failed_count) % 10 == 0:
                self.update_state(
                    state="PROGRESS",
                    meta={
                        "current": 3,
                        "total": 4,
                        "status": f"Processed {archived_count + failed_count}/{total_recordings}",
                    },
                )

        self.update_state(state="PROGRESS", meta={"current": 4, "total": 4, "status": "Archival completed"})

        logger.info(f"Archived {archived_count} recordings, {failed_count} failed")

        return {
            "total_recordings": total_recordings,
            "archived_count": archived_count,
            "failed_count": failed_count,
            "cutoff_date": cutoff_date.isoformat(),
            "completed_at": timezone.now().isoformat(),
        }

    except Exception as e:
        logger.error(f"Failed to archive old recordings: {e}")
        raise


@shared_task(name="cleanup_expired_sessions")
def cleanup_expired_sessions():
    """Clean up expired Django sessions and task execution records."""
    try:
        from django.contrib.sessions.models import Session

        from .models import TaskExecution

        # Clean up expired sessions
        Session.objects.filter(expire_date__lt=timezone.now()).delete()

        # Clean up old task execution records (keep last 30 days)
        cutoff_date = timezone.now() - timedelta(days=30)
        old_executions = TaskExecution.objects.filter(created_at__lt=cutoff_date)
        deleted_count, _ = old_executions.delete()

        logger.info(f"Cleaned up expired sessions and {deleted_count} old task executions")

        return {
            "task_executions_deleted": deleted_count,
            "cutoff_date": cutoff_date.isoformat(),
            "completed_at": timezone.now().isoformat(),
        }

    except Exception as e:
        logger.error(f"Failed to cleanup expired sessions: {e}")
        raise


# ================================
# EMAIL AND NOTIFICATIONS
# ================================


@shared_task(bind=True, name="send_report_email")
def send_report_email(self, report_data: Dict, recipients: List[str], report_type: str = "daily"):
    """Send analytics report via email.

    Args:
        report_data: Report data dictionary
        recipients: List of email addresses
        report_type: Type of report (daily, weekly, monthly)

    Returns:
        Dictionary with email sending results

    """
    try:
        self.update_state(state="PROGRESS", meta={"current": 1, "total": 3, "status": "Preparing email content"})

        # Prepare email content
        subject = f"Call Center {report_type.title()} Report"

        if report_type == "daily":
            subject += f" - {report_data.get('date', 'Unknown Date')}"
        elif report_type == "weekly":
            subject += f" - Week of {report_data.get('week_start', 'Unknown Week')}"
        elif report_type == "monthly":
            subject += f" - {report_data.get('month_name', 'Unknown Month')} {report_data.get('year', '')}"

        # Render email template
        context = {
            "report_data": report_data,
            "report_type": report_type,
            "generated_at": report_data.get("generated_at"),
        }

        html_content = render_to_string(f"django_twilio_call/emails/{report_type}_report.html", context)
        text_content = render_to_string(f"django_twilio_call/emails/{report_type}_report.txt", context)

        self.update_state(state="PROGRESS", meta={"current": 2, "total": 3, "status": "Sending email"})

        # Send email
        sent_count = 0
        failed_recipients = []

        for recipient in recipients:
            try:
                send_mail(
                    subject=subject,
                    message=text_content,
                    from_email=getattr(settings, "DEFAULT_FROM_EMAIL", "noreply@example.com"),
                    recipient_list=[recipient],
                    html_message=html_content,
                    fail_silently=False,
                )
                sent_count += 1
            except Exception as e:
                logger.error(f"Failed to send report email to {recipient}: {e}")
                failed_recipients.append(recipient)

        self.update_state(state="PROGRESS", meta={"current": 3, "total": 3, "status": "Email sending completed"})

        logger.info(f"Sent {report_type} report email to {sent_count} recipients")

        return {
            "report_type": report_type,
            "total_recipients": len(recipients),
            "sent_count": sent_count,
            "failed_recipients": failed_recipients,
            "sent_at": timezone.now().isoformat(),
        }

    except Exception as e:
        logger.error(f"Failed to send report email: {e}")
        raise


@shared_task(name="send_critical_alert")
def send_critical_alert(task_name: str, task_id: str, error: str):
    """Send critical alert for important task failures.

    Args:
        task_name: Name of the failed task
        task_id: Task ID
        error: Error message

    Returns:
        Dictionary with alert results

    """
    try:
        alert_recipients = getattr(settings, "CRITICAL_ALERT_RECIPIENTS", [])

        if not alert_recipients:
            logger.warning("No critical alert recipients configured")
            return {"message": "No alert recipients configured"}

        subject = f"CRITICAL: Task Failure - {task_name}"
        message = f"""
        A critical task has failed in the django-twilio-call system:

        Task Name: {task_name}
        Task ID: {task_id}
        Error: {error}
        Time: {timezone.now().isoformat()}

        Please investigate immediately.
        """

        sent_count = 0
        for recipient in alert_recipients:
            try:
                send_mail(
                    subject=subject,
                    message=message,
                    from_email=getattr(settings, "DEFAULT_FROM_EMAIL", "noreply@example.com"),
                    recipient_list=[recipient],
                    fail_silently=False,
                )
                sent_count += 1
            except Exception as e:
                logger.error(f"Failed to send critical alert to {recipient}: {e}")

        logger.info(f"Sent critical alert for task {task_name} to {sent_count} recipients")

        return {
            "task_name": task_name,
            "task_id": task_id,
            "sent_count": sent_count,
            "sent_at": timezone.now().isoformat(),
        }

    except Exception as e:
        logger.error(f"Failed to send critical alert: {e}")
        raise


# ================================
# WEBHOOK AND INTEGRATION TASKS
# ================================


@shared_task(bind=True, name="process_webhook_callback")
def process_webhook_callback(self, webhook_data: Dict, webhook_type: str):
    """Process incoming webhook callback from Twilio.

    Args:
        webhook_data: Webhook payload data
        webhook_type: Type of webhook (call-status, recording, etc.)

    Returns:
        Dictionary with processing results

    """
    try:
        from .models import WebhookLog

        # Log the webhook
        webhook_log = WebhookLog.objects.create(
            webhook_type=webhook_type,
            url=webhook_data.get("webhook_url", "unknown"),
            payload=webhook_data,
            status=WebhookLog.Status.PENDING,
        )

        self.update_state(
            state="PROGRESS", meta={"current": 1, "total": 3, "status": f"Processing {webhook_type} webhook"}
        )

        # Process based on webhook type
        if webhook_type == "call-status":
            result = _process_call_status_webhook(webhook_data)
        elif webhook_type == "recording":
            result = _process_recording_webhook(webhook_data)
        elif webhook_type == "transcription":
            result = _process_transcription_webhook(webhook_data)
        else:
            raise ValueError(f"Unknown webhook type: {webhook_type}")

        self.update_state(state="PROGRESS", meta={"current": 2, "total": 3, "status": "Updating related records"})

        # Update webhook log
        webhook_log.status = WebhookLog.Status.DELIVERED
        webhook_log.delivered_at = timezone.now()
        webhook_log.save(update_fields=["status", "delivered_at"])

        self.update_state(state="PROGRESS", meta={"current": 3, "total": 3, "status": "Webhook processing completed"})

        logger.info(f"Successfully processed {webhook_type} webhook")

        return {
            "webhook_type": webhook_type,
            "webhook_log_id": webhook_log.id,
            "result": result,
            "processed_at": timezone.now().isoformat(),
        }

    except Exception as e:
        # Update webhook log with failure
        if "webhook_log" in locals():
            webhook_log.status = WebhookLog.Status.FAILED
            webhook_log.error_message = str(e)
            webhook_log.save(update_fields=["status", "error_message"])

        logger.error(f"Failed to process {webhook_type} webhook: {e}")
        raise


@shared_task(bind=True, name="retry_failed_webhook")
def retry_failed_webhook(self, webhook_log_id: int):
    """Retry a failed webhook delivery.

    Args:
        webhook_log_id: WebhookLog ID to retry

    Returns:
        Dictionary with retry results

    """
    try:
        from .models import WebhookLog

        webhook_log = WebhookLog.objects.get(id=webhook_log_id)

        # Check if retry is needed
        if webhook_log.status == WebhookLog.Status.DELIVERED:
            return {"message": "Webhook already delivered"}

        if webhook_log.retry_count >= 3:
            webhook_log.status = WebhookLog.Status.ABANDONED
            webhook_log.abandoned_at = timezone.now()
            webhook_log.save(update_fields=["status", "abandoned_at"])
            return {"message": "Max retries exceeded, webhook abandoned"}

        self.update_state(state="PROGRESS", meta={"current": 1, "total": 2, "status": "Retrying webhook delivery"})

        # Increment retry count
        webhook_log.retry_count += 1
        webhook_log.status = WebhookLog.Status.RETRYING
        webhook_log.save(update_fields=["retry_count", "status"])

        # Retry processing
        result = process_webhook_callback.delay(webhook_data=webhook_log.payload, webhook_type=webhook_log.webhook_type)

        self.update_state(state="PROGRESS", meta={"current": 2, "total": 2, "status": "Retry initiated"})

        logger.info(f"Retrying webhook {webhook_log.id}, attempt {webhook_log.retry_count}")

        return {
            "webhook_log_id": webhook_log_id,
            "retry_count": webhook_log.retry_count,
            "retry_task_id": result.id,
            "retried_at": timezone.now().isoformat(),
        }

    except WebhookLog.DoesNotExist:
        logger.error(f"WebhookLog {webhook_log_id} not found")
        raise
    except Exception as e:
        logger.error(f"Failed to retry webhook {webhook_log_id}: {e}")
        raise


@shared_task(name="check_failed_webhooks")
def check_failed_webhooks():
    """Check for failed webhooks that need retry."""
    try:
        from .models import WebhookLog

        # Find failed webhooks that are ready for retry
        now = timezone.now()
        failed_webhooks = WebhookLog.objects.filter(
            status__in=[WebhookLog.Status.FAILED, WebhookLog.Status.RETRYING], retry_count__lt=3, next_retry_at__lte=now
        )

        if not failed_webhooks.exists():
            return {"message": "No failed webhooks to retry"}

        # Schedule retries
        retry_tasks = []
        for webhook_log in failed_webhooks:
            # Calculate next retry time with exponential backoff
            next_retry_delay = 60 * (2**webhook_log.retry_count)  # 1, 2, 4 minutes
            webhook_log.next_retry_at = now + timedelta(seconds=next_retry_delay)
            webhook_log.save(update_fields=["next_retry_at"])

            # Schedule retry
            retry_task = retry_failed_webhook.delay(webhook_log.id)
            retry_tasks.append(retry_task.id)

        logger.info(f"Scheduled retries for {len(retry_tasks)} failed webhooks")

        return {
            "retries_scheduled": len(retry_tasks),
            "retry_task_ids": retry_tasks,
            "scheduled_at": now.isoformat(),
        }

    except Exception as e:
        logger.error(f"Failed to check failed webhooks: {e}")
        raise


# ================================
# DATA EXPORT TASKS
# ================================


@shared_task(bind=True, name="export_call_data")
def export_call_data(self, filters: Dict, format: str = "csv", user_id: Optional[int] = None):
    """Export call data based on filters.

    Args:
        filters: Dictionary of filters to apply
        format: Export format (csv, excel, json)
        user_id: User ID requesting the export

    Returns:
        Dictionary with export results and download information

    """
    try:
        from .models import Call

        # Parse filters
        start_date = filters.get("start_date")
        end_date = filters.get("end_date")
        queue_ids = filters.get("queue_ids", [])
        agent_ids = filters.get("agent_ids", [])
        status_list = filters.get("status", [])

        self.update_state(state="PROGRESS", meta={"current": 1, "total": 5, "status": "Building query with filters"})

        # Build query
        queryset = Call.objects.select_related("agent__user", "queue", "phone_number_used").prefetch_related(
            "recordings", "logs"
        )

        if start_date:
            queryset = queryset.filter(created_at__gte=start_date)
        if end_date:
            queryset = queryset.filter(created_at__lte=end_date)
        if queue_ids:
            queryset = queryset.filter(queue_id__in=queue_ids)
        if agent_ids:
            queryset = queryset.filter(agent_id__in=agent_ids)
        if status_list:
            queryset = queryset.filter(status__in=status_list)

        total_calls = queryset.count()

        self.update_state(
            state="PROGRESS",
            meta={"current": 2, "total": 5, "status": f"Exporting {total_calls} calls in {format} format"},
        )

        # Export based on format
        if format.lower() == "csv":
            export_data = _export_calls_to_csv(queryset, self)
        elif format.lower() == "excel":
            export_data = _export_calls_to_excel(queryset, self)
        elif format.lower() == "json":
            export_data = _export_calls_to_json(queryset, self)
        else:
            raise ValueError(f"Unsupported export format: {format}")

        self.update_state(state="PROGRESS", meta={"current": 4, "total": 5, "status": "Preparing download"})

        # Generate download filename
        timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
        filename = f"call_export_{timestamp}.{format.lower()}"

        # Save to cache or file storage for download
        cache_key = f"export_{self.request.id}"
        cache.set(
            cache_key,
            {
                "data": export_data,
                "filename": filename,
                "format": format,
                "total_calls": total_calls,
            },
            3600,
        )  # Cache for 1 hour

        self.update_state(state="PROGRESS", meta={"current": 5, "total": 5, "status": "Export completed"})

        logger.info(f"Exported {total_calls} calls in {format} format")

        return {
            "total_calls": total_calls,
            "format": format,
            "filename": filename,
            "cache_key": cache_key,
            "download_expires_at": (timezone.now() + timedelta(hours=1)).isoformat(),
            "exported_at": timezone.now().isoformat(),
        }

    except Exception as e:
        logger.error(f"Failed to export call data: {e}")
        raise


# ================================
# SYSTEM MONITORING TASKS
# ================================


@shared_task(name="system_health_check")
def system_health_check():
    """Perform comprehensive system health check."""
    try:
        from .monitoring import get_system_status

        # Get comprehensive system status
        system_status = get_system_status()

        # Check for critical issues
        critical_issues = []

        # Check for high failure rates
        if system_status.get("health", {}).get("recent_success_rate", 100) < 90:
            critical_issues.append("High task failure rate detected")

        # Check for stuck tasks
        slow_tasks = system_status.get("slow_tasks", [])
        if len(slow_tasks) > 5:
            critical_issues.append(f"{len(slow_tasks)} slow-running tasks detected")

        # Check queue backlogs
        for queue_name, queue_data in system_status.get("queues", {}).items():
            if queue_data.get("active_tasks", 0) > 100:
                critical_issues.append(f"Queue {queue_name} has high backlog")

        # Send alerts if critical issues found
        if critical_issues:
            send_critical_alert.delay(
                task_name="system_health_check", task_id="health_monitor", error="; ".join(critical_issues)
            )

        # Cache health status
        cache.set(
            "system_health_status",
            {
                "status": "critical" if critical_issues else "healthy",
                "issues": critical_issues,
                "checked_at": timezone.now().isoformat(),
                "full_status": system_status,
            },
            300,
        )  # Cache for 5 minutes

        logger.info(f"System health check completed, {len(critical_issues)} critical issues found")

        return {
            "status": "critical" if critical_issues else "healthy",
            "critical_issues": critical_issues,
            "checked_at": timezone.now().isoformat(),
        }

    except Exception as e:
        logger.error(f"System health check failed: {e}")
        raise


@shared_task(name="update_all_agent_metrics")
def update_all_agent_metrics():
    """Update metrics for all active agents."""
    try:
        from .models import Agent

        active_agents = Agent.objects.filter(is_active=True)
        total_agents = active_agents.count()

        if total_agents == 0:
            return {"message": "No active agents to update"}

        # Create group of metric calculation tasks
        job = group(calculate_agent_metrics.s(agent.id) for agent in active_agents)

        result = job.apply_async()

        logger.info(f"Scheduled metric updates for {total_agents} agents")

        return {
            "total_agents": total_agents,
            "group_id": result.id,
            "scheduled_at": timezone.now().isoformat(),
        }

    except Exception as e:
        logger.error(f"Failed to update agent metrics: {e}")
        raise


@shared_task(name="optimize_queue_routing")
def optimize_queue_routing():
    """Analyze and optimize queue routing strategies."""
    try:
        from .models import Queue
        from .services.analytics_service import analytics_service

        # Get queue performance data
        yesterday = timezone.now() - timedelta(days=1)

        optimization_results = []

        for queue in Queue.objects.filter(is_active=True):
            # Get queue analytics
            queue_analytics = analytics_service.get_queue_analytics(
                queue_id=queue.id, start_date=yesterday, end_date=timezone.now()
            )

            if not queue_analytics["queues"]:
                continue

            queue_data = queue_analytics["queues"][0]

            # Analyze performance and suggest optimizations
            suggestions = []

            # Check service level
            if queue_data["performance"]["service_level"] < 80:
                suggestions.append("Consider adding more agents or adjusting routing strategy")

            # Check abandonment rate
            if queue_data["performance"]["abandonment_rate"] > 10:
                suggestions.append("High abandonment rate - review queue timeout settings")

            # Check agent utilization
            if queue_data["agents"]["utilization"] > 90:
                suggestions.append("Agent utilization very high - consider load balancing")

            optimization_results.append(
                {
                    "queue_id": queue.id,
                    "queue_name": queue.name,
                    "current_strategy": queue.routing_strategy,
                    "performance": queue_data["performance"],
                    "suggestions": suggestions,
                }
            )

        # Cache optimization results
        cache.set(
            "queue_optimization_results",
            {
                "results": optimization_results,
                "analyzed_at": timezone.now().isoformat(),
            },
            1800,
        )  # Cache for 30 minutes

        logger.info(f"Analyzed {len(optimization_results)} queues for optimization")

        return {
            "queues_analyzed": len(optimization_results),
            "analyzed_at": timezone.now().isoformat(),
        }

    except Exception as e:
        logger.error(f"Failed to optimize queue routing: {e}")
        raise


# ================================
# HELPER FUNCTIONS
# ================================


def _calculate_percentage_change(old_value: float, new_value: float) -> float:
    """Calculate percentage change between two values."""
    if old_value == 0:
        return 100.0 if new_value > 0 else 0.0

    return round(((new_value - old_value) / old_value) * 100, 2)


def _process_call_status_webhook(webhook_data: Dict) -> Dict:
    """Process call status webhook."""
    from .services.call_service import call_service

    call_sid = webhook_data.get("CallSid")
    call_status = webhook_data.get("CallStatus")

    if not call_sid:
        raise ValueError("CallSid missing from webhook data")

    # Update call status
    result = call_service.update_call_status(call_sid, call_status, webhook_data)

    return {"call_sid": call_sid, "status": call_status, "updated": True}


def _process_recording_webhook(webhook_data: Dict) -> Dict:
    """Process recording webhook."""
    from .services.recording_service import recording_service

    recording_sid = webhook_data.get("RecordingSid")

    if not recording_sid:
        raise ValueError("RecordingSid missing from webhook data")

    # Process recording
    recording = recording_service.process_recording_callback(webhook_data)

    # Schedule async processing
    process_call_recording.delay(recording.call.id, webhook_data)

    return {"recording_sid": recording_sid, "processed": True}


def _process_transcription_webhook(webhook_data: Dict) -> Dict:
    """Process transcription webhook."""
    from .models import CallRecording

    recording_sid = webhook_data.get("RecordingSid")
    transcription_text = webhook_data.get("TranscriptionText", "")
    transcription_status = webhook_data.get("TranscriptionStatus", "completed")

    if not recording_sid:
        raise ValueError("RecordingSid missing from webhook data")

    # Update recording with transcription
    recording = CallRecording.objects.get(twilio_sid=recording_sid)
    recording.transcription = transcription_text
    recording.transcription_status = transcription_status
    recording.save(update_fields=["transcription", "transcription_status"])

    return {"recording_sid": recording_sid, "transcription_status": transcription_status}


def _export_calls_to_csv(queryset, task) -> str:
    """Export calls to CSV format."""
    import csv
    import io

    output = io.StringIO()
    writer = csv.writer(output)

    # Write header
    writer.writerow(
        [
            "Call ID",
            "Twilio SID",
            "From Number",
            "To Number",
            "Direction",
            "Status",
            "Agent",
            "Queue",
            "Duration",
            "Queue Time",
            "Created At",
            "Answered At",
            "End Time",
            "Price",
            "Caller Name",
        ]
    )

    # Write data
    for i, call in enumerate(queryset.iterator(chunk_size=1000)):
        writer.writerow(
            [
                call.id,
                call.twilio_sid,
                call.from_number,
                call.to_number,
                call.direction,
                call.status,
                call.agent.user.get_full_name() if call.agent else "",
                call.queue.name if call.queue else "",
                call.duration,
                call.queue_time,
                call.created_at.isoformat(),
                call.answered_at.isoformat() if call.answered_at else "",
                call.end_time.isoformat() if call.end_time else "",
                str(call.price) if call.price else "",
                call.caller_name,
            ]
        )

        # Update progress every 100 calls
        if i % 100 == 0:
            task.update_state(state="PROGRESS", meta={"current": 3, "total": 5, "status": f"Exported {i} calls"})

    return output.getvalue()


def _export_calls_to_excel(queryset, task) -> bytes:
    """Export calls to Excel format."""
    try:
        from io import BytesIO

        import openpyxl
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Call Data"

        # Write header
        headers = [
            "Call ID",
            "Twilio SID",
            "From Number",
            "To Number",
            "Direction",
            "Status",
            "Agent",
            "Queue",
            "Duration",
            "Queue Time",
            "Created At",
            "Answered At",
            "End Time",
            "Price",
            "Caller Name",
        ]

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # Write data
        for i, call in enumerate(queryset.iterator(chunk_size=1000), 2):
            ws.cell(row=i, column=1, value=call.id)
            ws.cell(row=i, column=2, value=call.twilio_sid)
            ws.cell(row=i, column=3, value=call.from_number)
            ws.cell(row=i, column=4, value=call.to_number)
            ws.cell(row=i, column=5, value=call.direction)
            ws.cell(row=i, column=6, value=call.status)
            ws.cell(row=i, column=7, value=call.agent.user.get_full_name() if call.agent else "")
            ws.cell(row=i, column=8, value=call.queue.name if call.queue else "")
            ws.cell(row=i, column=9, value=call.duration)
            ws.cell(row=i, column=10, value=call.queue_time)
            ws.cell(row=i, column=11, value=call.created_at)
            ws.cell(row=i, column=12, value=call.answered_at)
            ws.cell(row=i, column=13, value=call.end_time)
            ws.cell(row=i, column=14, value=float(call.price) if call.price else 0)
            ws.cell(row=i, column=15, value=call.caller_name)

            # Update progress every 100 calls
            if (i - 1) % 100 == 0:
                task.update_state(
                    state="PROGRESS", meta={"current": 3, "total": 5, "status": f"Exported {i - 1} calls"}
                )

        # Save to bytes
        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    except ImportError:
        raise ValueError("openpyxl is required for Excel export")


def _export_calls_to_json(queryset, task) -> str:
    """Export calls to JSON format."""
    import json

    calls_data = []

    for i, call in enumerate(queryset.iterator(chunk_size=1000)):
        call_data = {
            "id": call.id,
            "twilio_sid": call.twilio_sid,
            "from_number": call.from_number,
            "to_number": call.to_number,
            "direction": call.direction,
            "status": call.status,
            "agent": call.agent.user.get_full_name() if call.agent else None,
            "queue": call.queue.name if call.queue else None,
            "duration": call.duration,
            "queue_time": call.queue_time,
            "created_at": call.created_at.isoformat(),
            "answered_at": call.answered_at.isoformat() if call.answered_at else None,
            "end_time": call.end_time.isoformat() if call.end_time else None,
            "price": float(call.price) if call.price else None,
            "caller_name": call.caller_name,
            "metadata": call.metadata,
        }
        calls_data.append(call_data)

        # Update progress every 100 calls
        if i % 100 == 0:
            task.update_state(state="PROGRESS", meta={"current": 3, "total": 5, "status": f"Exported {i} calls"})

    return json.dumps(
        {
            "calls": calls_data,
            "total_count": len(calls_data),
            "exported_at": timezone.now().isoformat(),
        },
        indent=2,
    )
