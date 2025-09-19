# encoding: utf-8

from openpyxl import load_workbook


def excel_read(xlsx, sheet):
    wb = load_workbook(rf'{xlsx}')
    sheet = wb[rf'{sheet}']
    result = []

    for row in range(1, sheet.max_row + 1):

        single_line = []
        for column in range(1, sheet.max_column + 1):
            single_line.append(sheet.cell(row, column).value)

        result.append(single_line)

    return result, sheet.max_column
